"""
Parseur de mapping comptable et de balances N-1.

Sources de mapping :
  1. FM existant du client (onglet "Balance N Vs N-1")
  2. Fichier mapping_pcg.yaml (config générique du cabinet)

Sources de balance N-1 :
  3. FEC N-1 (.txt) — construit la balance par agrégation
  4. Balance Excel simple (.xlsx) — colonnes CompteNum, CompteLib, Solde (en €)
"""

import logging
from pathlib import Path
from typing import Dict, Optional, Tuple, Union

import openpyxl
import pandas as pd
import yaml

logger = logging.getLogger(__name__)

# Type canonique du mapping :
# {compte_num_str: {cycle, etatfi_n, etatfi_n1, compta_n, compta_n1,
#                   etatfi (alias de etatfi_n), compta (alias de compta_n), ref}}
MappingCompte = Dict[str, Dict[str, str]]


def from_fm(path: Union[str, Path]) -> MappingCompte:
    """
    Extrait le mapping depuis l'onglet "Balance N Vs N-1" d'un FM existant.

    Structure de l'onglet (headers en ligne 8, données à partir de ligne 10) :
      col B (index 1) : CompteNum
      col H (index 7) : Ref cycle   (ex: "C Propres0")
      col J (index 9) : Cycle       (ex: "C Propres", "A")
      col K (index 10): EtatFi N    (ex: "Capital social")
      col L (index 11): EtatFi N-1  (peut différer de EtatFi N — reclassement)
      col M (index 12): ComptaN     (ex: "Passif", "Actif")
      col N (index 13): ComptaN-1   (peut différer de ComptaN — reclassement)

    Le dict retourné contient les clés distinctes etatfi_n / etatfi_n1 /
    compta_n / compta_n1, plus les aliases historiques etatfi (= etatfi_n)
    et compta (= compta_n) pour ne casser aucun consommateur existant.
    """
    chemin = Path(path)
    if not chemin.exists():
        raise FileNotFoundError(f"FM introuvable : {chemin}")

    wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)

    if "Balance N Vs N-1" not in wb.sheetnames:
        raise ValueError(f"Onglet 'Balance N Vs N-1' absent de {chemin.name}")

    ws = wb["Balance N Vs N-1"]
    mapping: MappingCompte = {}

    for row in ws.iter_rows(min_row=10, values_only=True):
        compte_raw = row[1]
        if compte_raw is None:
            continue
        try:
            compte_num = str(int(float(compte_raw)))
        except (ValueError, TypeError):
            continue

        cycle     = str(row[9]).strip()  if row[9]  is not None else ""
        etatfi_n  = str(row[10]).strip() if row[10] is not None else ""
        etatfi_n1 = str(row[11]).strip() if row[11] is not None else ""
        compta_n  = str(row[12]).strip() if row[12] is not None else ""
        compta_n1 = str(row[13]).strip() if row[13] is not None else ""
        ref       = str(row[7]).strip()  if row[7]  is not None else ""

        if not cycle:
            continue

        mapping[compte_num] = {
            "cycle":     cycle,
            "etatfi_n":  etatfi_n,
            "etatfi_n1": etatfi_n1,
            "compta_n":  compta_n,
            "compta_n1": compta_n1,
            # Aliases historiques (consommateurs existants)
            "etatfi":    etatfi_n,
            "compta":    compta_n,
            "ref":       ref,
        }

    logger.info("FM '%s' : %d comptes chargés", chemin.name, len(mapping))
    return mapping


def from_pcg_config(path: Union[str, Path]) -> dict:
    """
    Charge et structure le fichier mapping_pcg.yaml.

    Retourne un dict avec les clés :
      prefixes        : dict {prefixe_str: cycle}  — toutes classes fusionnées
      surcharges      : dict {prefixe_str: cycle}  — surcharges cabinet
      passif_prefixes : set  {prefixe_str}          — préfixes classe 4 → Passif
      ordre_cycles    : list[str]
      noms_cycles     : dict {code: nom_long}
      seuils          : dict
      templates       : dict {nom_template: cycle}
      liasse_fiscale  : dict — section brute du YAML (préfixes des états
                        financiers Bilan/Tréso/AACE/EBIT/P&L), validée par
                        src.engine.liasse_fiscale_loader.load_liasse_fiscale
    """
    chemin = Path(path)
    if not chemin.exists():
        raise FileNotFoundError(f"Config PCG introuvable : {chemin}")

    with chemin.open(encoding="utf-8") as f:
        config = yaml.safe_load(f)

    # Fusionner toutes les classes (classe_1 … classe_7) en un seul dict plat
    prefixes: Dict[str, str] = {}
    for cle, contenu in config.items():
        if cle.startswith("classe_") and isinstance(contenu, dict):
            for prefixe, cycle in contenu.items():
                prefixes[str(prefixe)] = str(cycle)

    surcharges = {
        str(k): str(v)
        for k, v in (config.get("surcharges_cabinet") or {}).items()
    }

    passif_prefixes = set(
        str(p)
        for p in (config.get("classification_bilan", {}).get("passif") or [])
    )

    templates = {
        str(k): str(v)
        for k, v in (config.get("templates") or {}).items()
    }

    logger.info(
        "PCG config '%s' : %d préfixes, %d surcharges, %d préfixes passif",
        chemin.name, len(prefixes), len(surcharges), len(passif_prefixes),
    )

    return {
        "prefixes":        prefixes,
        "surcharges":      surcharges,
        "passif_prefixes": passif_prefixes,
        "ordre_cycles":    config.get("ordre_cycles", []),
        "noms_cycles":     config.get("noms_cycles", {}),
        "seuils":          config.get("seuils", {}),
        "templates":       templates,
        "liasse_fiscale":  config.get("liasse_fiscale") or {},
    }


# ---------------------------------------------------------------------------
# Sources de balance N-1
# ---------------------------------------------------------------------------

BalanceN1Dict = Dict[str, Dict[str, Union[str, float]]]


def from_balance_excel(path: Union[str, Path]) -> BalanceN1Dict:
    """
    Lit un fichier Excel balance simple avec colonnes CompteNum, CompteLib, Solde (en €).

    La recherche des colonnes est insensible à la casse et aux espaces.
    Le solde est converti de € en K€ (÷ 1000).

    Retourne
    --------
    dict {compte_num_str: {"libelle": str, "solde_ke": float}}

    Lève
    ----
    FileNotFoundError
        Si le fichier est introuvable.
    ValueError
        Si les colonnes obligatoires (CompteNum, Solde) sont absentes.
    """
    chemin = Path(path)
    if not chemin.exists():
        raise FileNotFoundError(f"Balance Excel introuvable : {chemin}")

    df = pd.read_excel(chemin, dtype=str)

    # Normaliser les noms de colonnes pour la recherche flexible
    col_map: Dict[str, str] = {str(c).strip().lower(): c for c in df.columns}

    def _find_col(candidates: list) -> Optional[str]:
        for candidate in candidates:
            if candidate.lower() in col_map:
                return col_map[candidate.lower()]
        return None

    col_num  = _find_col(["comptenum", "compte_num", "compte num", "num", "numero"])
    col_lib  = _find_col(["comptelib", "compte_lib", "libelle", "lib", "intitule"])
    col_sold = _find_col(["solde", "solde_n", "solde n", "balance", "montant"])

    # Fallback positionnel : si pas d'en-têtes reconnus et exactement 3 colonnes,
    # on suppose col0=CompteNum, col1=CompteLib, col2=Solde
    cols = list(df.columns)
    if col_num is None and len(cols) >= 3:
        logger.warning(
            "Balance Excel '%s' : en-têtes non reconnus — "
            "assignation positionnelle (col1=CompteNum, col2=CompteLib, col3=Solde)",
            chemin.name,
        )
        col_num  = cols[0]
        col_lib  = cols[1]
        col_sold = cols[2]

    if col_num is None:
        raise ValueError(
            f"Colonne CompteNum introuvable dans {chemin.name}. "
            f"Colonnes disponibles : {list(df.columns)}"
        )
    if col_sold is None:
        raise ValueError(
            f"Colonne Solde introuvable dans {chemin.name}. "
            f"Colonnes disponibles : {list(df.columns)}"
        )

    balance: BalanceN1Dict = {}
    lignes_ignorees = 0

    for _, row in df.iterrows():
        compte_raw = row[col_num]
        if pd.isna(compte_raw) or str(compte_raw).strip() == "":
            lignes_ignorees += 1
            continue
        try:
            compte_num = str(int(float(str(compte_raw).strip())))
        except (ValueError, TypeError):
            lignes_ignorees += 1
            continue

        libelle = str(row[col_lib]).strip() if col_lib and not pd.isna(row[col_lib]) else ""

        solde_raw = row[col_sold]
        try:
            solde_eur = float(str(solde_raw).strip().replace(",", ".").replace(" ", "")) if not pd.isna(solde_raw) else 0.0
        except (ValueError, TypeError):
            solde_eur = 0.0

        balance[compte_num] = {
            "libelle":  libelle,
            "solde_ke": round(solde_eur / 1000, 3),
        }

    logger.info(
        "Balance Excel '%s' : %d comptes chargés (%d lignes ignorées)",
        chemin.name, len(balance), lignes_ignorees,
    )
    return balance


def from_fec_n1(path: Union[str, Path]) -> BalanceN1Dict:
    """
    Lit un FEC N-1 et construit la balance agrégée par compte.

    Utilise fec_parser.parse() pour lire le FEC, puis agrège par CompteNum
    pour calculer le solde (Débit - Crédit) et convertit en K€.

    Retourne
    --------
    dict {compte_num_str: {"libelle": str, "solde_ke": float}}

    Lève
    ----
    FileNotFoundError
        Si le fichier est introuvable.
    ValueError
        Si le FEC est invalide (colonnes manquantes, encodage inconnu).
    """
    from src.parsers.fec_parser import parse as _parse_fec

    chemin = Path(path)
    if not chemin.exists():
        raise FileNotFoundError(f"FEC N-1 introuvable : {chemin}")

    logger.info("Chargement du FEC N-1 : %s", chemin.name)
    df = _parse_fec(chemin)

    # Agrégation par compte : prendre le dernier libellé rencontré (le plus récent)
    agg = (
        df.groupby("CompteNum", as_index=False)
        .agg(
            CompteLib=("CompteLib", "last"),
            Solde=("Solde", "sum"),
        )
    )

    balance: BalanceN1Dict = {
        str(row["CompteNum"]): {
            "libelle":  str(row["CompteLib"]),
            "solde_ke": round(float(row["Solde"]) / 1000, 3),
        }
        for _, row in agg.iterrows()
    }

    logger.info(
        "FEC N-1 '%s' : %d comptes agrégés depuis %d lignes",
        chemin.name, len(balance), len(df),
    )
    return balance


def detect_balance_sheet(path: Union[str, Path]) -> Tuple[str, str]:
    """
    Détecte automatiquement la feuille contenant la balance dans un classeur Excel.

    Retourne (nom_feuille, mode) où mode est :
      - "fm"      si c'est un FM complet (onglet nommé avec "balance")
      - "balance" si c'est une balance simple (détection par contenu)

    Lève ValueError si aucune feuille n'est détectée.
    """
    chemin = Path(path)
    wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
    sheetnames = wb.sheetnames

    # Priorité 1 — nom exact
    if "Balance N Vs N-1" in sheetnames:
        wb.close()
        logger.info("Feuille balance détectée (mode fm) : 'Balance N Vs N-1'")
        return "Balance N Vs N-1", "fm"

    # Priorité 2 — nom contient "balance" (insensible à la casse)
    for nom in sheetnames:
        if "balance" in nom.lower():
            wb.close()
            logger.info("Feuille balance détectée (mode fm) : '%s'", nom)
            return nom, "fm"

    # Priorité 3 — nom contient "bal" (insensible à la casse)
    for nom in sheetnames:
        if "bal" in nom.lower():
            wb.close()
            logger.info("Feuille balance détectée (mode fm) : '%s'", nom)
            return nom, "fm"

    # Priorité 4 — scan du contenu : chercher "CompteNum", "Compte" ou "Num"
    _MOTS_CLES = {"comptenum", "compte", "num"}
    for nom in sheetnames:
        ws = wb[nom]
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            for cell in row:
                if isinstance(cell, str) and cell.strip().lower() in _MOTS_CLES:
                    wb.close()
                    logger.info(
                        "Feuille balance détectée (mode balance) : '%s'", nom
                    )
                    return nom, "balance"

    wb.close()
    raise ValueError(
        f"Aucune feuille balance détectée dans {path}. "
        f"Onglets disponibles : {sheetnames}"
    )
