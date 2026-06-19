"""
Générateur du fichier de travail Travail_{client}_{annee}.xlsx.

Onglets produits :
  1. FEC N         — FEC brut complet (18 colonnes obligatoires + Solde)
  2. Balance N     — Agrégat par compte : CompteNum, CompteLib, Débit, Crédit, Solde, Solde K€
  3. Balance N-1   — CompteNum, CompteLib, Solde K€
"""

import logging
from pathlib import Path
from typing import Optional, Union

import pandas as pd
from openpyxl import Workbook

from src.writers.styles import (
    remove_gridlines, set_col_widths,
    write_header_row, write_data_row,
    FONT_NORMAL, FONT_BOLD, FONT_SUBTITLE,
    FONT_TITLE, NUM_KE,
)

logger = logging.getLogger(__name__)

# Capacité maximale d'une feuille Excel (lignes). Les lignes 1-4 sont
# réservées au titre et aux en-têtes ; au-delà, le FEC est exporté en CSV.
_EXCEL_MAX_ROWS = 1_048_576
_FEC_MAX_DATA_ROWS = _EXCEL_MAX_ROWS - 4

# Colonnes FEC dans l'ordre réglementaire + Solde calculé
_COLONNES_FEC = [
    "JournalCode", "JournalLib", "EcritureNum", "EcritureDate",
    "CompteNum", "CompteLib", "CompAuxNum", "CompAuxLib",
    "PieceRef", "PieceDate", "EcritureLib",
    "Debit", "Credit",
    "EcritureLet", "DateLet", "ValidDate", "Montantdevise", "Idevise",
    "Solde",
]

# Largeurs des colonnes de l'onglet FEC (par lettre Excel)
_WIDTHS_FEC = {
    "A": 10, "B": 22, "C": 14, "D": 12,
    "E": 12, "F": 32, "G": 12, "H": 28,
    "I": 14, "J": 12, "K": 32,
    "L": 14, "M": 14,
    "N": 10, "O": 12, "P": 12, "Q": 14, "R": 10,
    "S": 14,   # Solde
}

# Largeurs des colonnes de l'onglet Balance N
_WIDTHS_BN = {
    "A": 12, "B": 38, "C": 16, "D": 16, "E": 16, "F": 14,
}

# Largeurs des colonnes de l'onglet Balance N-1
_WIDTHS_BN1 = {
    "A": 12, "B": 38, "C": 14,
}


# ---------------------------------------------------------------------------
# Onglet FEC N
# ---------------------------------------------------------------------------

def _ecrire_fec(
    ws,
    df_fec: pd.DataFrame,
    client: str,
    date_cloture: str,
    dossier: Path,
    annee: int,
) -> None:
    """
    Écrit le FEC brut complet dans la feuille ws.

    Performance : un FEC peut compter des centaines de milliers de lignes
    (≈ 6 M cellules). Seul l'en-tête est stylé ; les cellules de données sont
    écrites sans police ni bordure et via itertuples (≈ 10× plus rapide
    qu'iterrows). On évite ainsi le coût d'application d'un style sur des
    millions de cellules — Design A ne s'applique qu'aux états (FM/Balance),
    pas au copier-coller brut du FEC.

    Scalabilité : si le FEC dépasse la capacité d'une feuille Excel
    (1 048 576 lignes), il est exporté en CSV compagnon FEC_{client}_{annee}.csv
    et l'onglet ne contient qu'un renvoi vers ce fichier.
    """
    remove_gridlines(ws)
    set_col_widths(ws, _WIDTHS_FEC)

    # Titre (lignes 1-2)
    ws.cell(row=1, column=1, value=f"FEC — {client}").font = FONT_TITLE
    ws.cell(row=2, column=1, value=f"Exercice clos le {date_cloture}").font = FONT_SUBTITLE

    # En-têtes (ligne 4)
    colonnes_disponibles = [c for c in _COLONNES_FEC if c in df_fec.columns]
    headers = [(i + 1, col) for i, col in enumerate(colonnes_disponibles)]
    write_header_row(ws, row=4, headers=headers)

    df_export = df_fec[colonnes_disponibles]
    nb_lignes = len(df_export)
    logger.info("FEC N : écriture de %d lignes", nb_lignes)

    # Garde-fou capacité Excel : au-delà, export CSV compagnon + renvoi.
    if nb_lignes > _FEC_MAX_DATA_ROWS:
        csv_path = dossier / f"FEC_{client}_{annee}.csv"
        df_export.to_csv(csv_path, sep=";", index=False, encoding="utf-8-sig")
        logger.warning(
            "FEC trop volumineux pour Excel (%d lignes > %d) — export CSV "
            "compagnon : %s ; l'onglet 'FEC N' ne contient qu'un renvoi.",
            nb_lignes, _FEC_MAX_DATA_ROWS, csv_path.name,
        )
        ws.cell(
            row=5, column=1,
            value=(f"FEC trop volumineux pour Excel ({nb_lignes:,} lignes). "
                   f"Voir le fichier compagnon : {csv_path.name}"),
        ).font = FONT_NORMAL
        return

    # Index des colonnes numériques (format appliqué cellule par cellule mais
    # SANS police ni bordure — le coût dominant des styles est ainsi évité).
    cols_num = {
        j for j, col in enumerate(colonnes_disponibles)
        if col in ("Debit", "Credit", "Solde", "Montantdevise")
    }

    # NaN -> None vectorisé une seule fois, puis itertuples.
    records = (
        df_export.astype(object)
        .where(pd.notna(df_export), None)
        .itertuples(index=False, name=None)
    )
    row_excel = 4  # la première ligne de données sera la 5
    for rec in records:
        row_excel += 1
        for j, valeur in enumerate(rec):
            c = ws.cell(row=row_excel, column=j + 1, value=valeur)
            if j in cols_num and valeur is not None:
                c.number_format = "#,##0.00"


# ---------------------------------------------------------------------------
# Onglet Balance N
# ---------------------------------------------------------------------------

def _ecrire_balance_n(ws, balance: pd.DataFrame, client: str, date_cloture: str) -> None:
    """Écrit la balance N agrégée dans la feuille ws."""
    remove_gridlines(ws)
    set_col_widths(ws, _WIDTHS_BN)

    # Titre
    ws.cell(row=1, column=1, value=f"Balance N — {client}").font = FONT_TITLE
    ws.cell(row=2, column=1, value=f"Exercice clos le {date_cloture}").font = FONT_SUBTITLE

    # En-têtes (ligne 4)
    headers = [
        (1, "CompteNum"),
        (2, "CompteLib"),
        (3, "Débit"),
        (4, "Crédit"),
        (5, "Solde"),
        (6, "Solde K€"),
    ]
    write_header_row(ws, row=4, headers=headers)

    # Tri par CompteNum
    df_tri = balance.sort_values("CompteNum").reset_index(drop=True)

    # Colonnes nécessaires (avec fallback si absentes)
    col_debit  = "Debit"  if "Debit"  in df_tri.columns else None
    col_credit = "Credit" if "Credit" in df_tri.columns else None
    col_solde  = "Solde"  if "Solde"  in df_tri.columns else None
    col_ke     = "Solde_KE" if "Solde_KE" in df_tri.columns else None

    logger.info("Balance N : écriture de %d comptes", len(df_tri))

    for i, row_data in df_tri.iterrows():
        row_excel = 5 + i
        cells = [
            (1, str(row_data["CompteNum"])),
            (2, row_data["CompteLib"]),
            (3, float(row_data[col_debit])  if col_debit  else None, "#,##0.00"),
            (4, float(row_data[col_credit]) if col_credit else None, "#,##0.00"),
            (5, float(row_data[col_solde])  if col_solde  else None, "#,##0.00"),
            (6, float(row_data[col_ke])     if col_ke     else None, NUM_KE),
        ]
        write_data_row(ws, row=row_excel, cells=cells)

    # Ligne total
    row_total = 5 + len(df_tri)
    ws.cell(row=row_total, column=2, value="TOTAL").font = FONT_BOLD
    if col_ke:
        total_ke = df_tri[col_ke].sum()
        c = ws.cell(row=row_total, column=6, value=round(total_ke))
        c.number_format = NUM_KE
        c.font = FONT_BOLD


# ---------------------------------------------------------------------------
# Onglet Balance N-1
# ---------------------------------------------------------------------------

def _ecrire_balance_n1(
    ws,
    balance_n1: dict,
    client: str,
    date_cloture: str,
) -> None:
    """Écrit la balance N-1 (dict) dans la feuille ws."""
    remove_gridlines(ws)
    set_col_widths(ws, _WIDTHS_BN1)

    # Calcul de l'année N-1
    annee_n = pd.to_datetime(date_cloture, format="%d/%m/%Y").year
    date_n1 = date_cloture.replace(str(annee_n), str(annee_n - 1))

    # Titre
    ws.cell(row=1, column=1, value=f"Balance N-1 — {client}").font = FONT_TITLE
    ws.cell(row=2, column=1, value=f"Exercice clos le {date_n1}").font = FONT_SUBTITLE

    # En-têtes (ligne 4)
    headers = [
        (1, "CompteNum"),
        (2, "CompteLib"),
        (3, "Solde K€"),
    ]
    write_header_row(ws, row=4, headers=headers)

    # Tri par numéro de compte
    comptes_tries = sorted(balance_n1.items(), key=lambda x: str(x[0]))
    logger.info("Balance N-1 : écriture de %d comptes", len(comptes_tries))

    for i, (compte_num, infos) in enumerate(comptes_tries):
        row_excel = 5 + i
        solde_ke = infos.get("solde_ke", 0.0)
        cells = [
            (1, str(compte_num)),
            (2, infos.get("libelle", "")),
            (3, float(solde_ke) if solde_ke is not None else None, NUM_KE),
        ]
        write_data_row(ws, row=row_excel, cells=cells)


# ---------------------------------------------------------------------------
# Point d'entrée public
# ---------------------------------------------------------------------------

def write(
    df_fec: pd.DataFrame,
    balance: pd.DataFrame,
    balance_n1: Optional[dict],
    client: str,
    date_cloture: str,
    output_path: Union[str, Path],
) -> Path:
    """
    Génère le fichier de travail Travail_{client}_{annee}.xlsx.

    Paramètres
    ----------
    df_fec        : DataFrame brut du FEC (produit par fec_parser.parse())
    balance       : DataFrame de la balance N (produit par balance_builder.build())
    balance_n1    : dict {compte_num: {"libelle": str, "solde_ke": float}} ou None
    client        : nom du client (ex. "GILAC")
    date_cloture  : date au format "JJ/MM/AAAA" (ex. "31/12/2025")
    output_path   : dossier de sortie

    Retourne
    --------
    Path du fichier généré.
    """
    annee = pd.to_datetime(date_cloture, format="%d/%m/%Y").year
    dossier = Path(output_path)
    dossier.mkdir(parents=True, exist_ok=True)

    nom_fichier = f"Travail_{client}_{annee}.xlsx"
    chemin = dossier / nom_fichier

    wb = Workbook()
    # Supprimer la feuille par défaut créée par openpyxl
    wb.remove(wb.active)

    # --- Onglet 1 : FEC N ---
    ws_fec = wb.create_sheet("FEC N")
    _ecrire_fec(ws_fec, df_fec, client, date_cloture, dossier, annee)

    # --- Onglet 2 : Balance N ---
    ws_bn = wb.create_sheet("Balance N")
    _ecrire_balance_n(ws_bn, balance, client, date_cloture)

    # --- Onglet 3 : Balance N-1 ---
    if balance_n1:
        ws_bn1 = wb.create_sheet("Balance N-1")
        _ecrire_balance_n1(ws_bn1, balance_n1, client, date_cloture)
    else:
        logger.warning("Balance N-1 absente — onglet 'Balance N-1' non généré")

    wb.save(chemin)
    logger.info("Fichier de travail généré : %s", chemin)
    return chemin
