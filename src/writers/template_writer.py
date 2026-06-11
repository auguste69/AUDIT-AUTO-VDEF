"""
Injection des données client dans les templates de feuilles de travail.

Pour chaque template .xlsx dans le dossier source :
  1. Copie le fichier
  2. Renomme : 20XX_XX_YYY_  →  {annee}_12_{client}_
  3. Remplace dans toutes les cellules de toutes les feuilles :
       #NomClient    → nom du client
       #DateClôture  → date au format JJ/MM/AAAA
       #Date         → idem
  4. Supprime les quadrillages de toutes les feuilles
  5. Insère les feuilles maîtresses du FM comme NOUVEAUX ONGLETS, juste
     avant la feuille Synthèse (si fm_path et integration_templates fournis)
  6. Restaure les images/dessins/commentaires perdus par openpyxl
     (manipulation ZIP/XML directe — voir _restaurer_dessins)
  7. Compresse l'ensemble dans un ZIP

Les templates sans cycle reconnu dans la config sont ignorés avec un
avertissement.

Préservation des images (P1, révision 09/06/2026) : 9 des 13 templates
contiennent des images ou dessins et openpyxl les perd à la simple
ouverture+sauvegarde. Après la sauvegarde openpyxl, les parties
xl/drawings/*, xl/media/* et xl/comments*.xml du template SOURCE sont
réinjectées dans l'archive générée, et les références <drawing> /
<legacyDrawing> des feuilles sont reconstruites (correspondance par NOM
de feuille — les feuilles FM insérées décalent les index sheetN.xml).
"""

import io
import logging
import re
import zipfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union
from xml.etree import ElementTree

import openpyxl
import pandas as pd

from src.writers.styles import remove_gridlines
from src.writers.worksheet_copy import copy_worksheet

logger = logging.getLogger(__name__)

# Préfixe commun à tous les noms de fichiers template
_PREFIXE_TEMPLATE = "20XX_XX_YYY_"

# Placeholders reconnus et leur remplacement cible
_PLACEHOLDERS = ("#NomClient", "#DateClôture", "#Date", "#date de clôture")

# Namespaces OOXML
_NS_REL_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"
_NS_CT      = "http://schemas.openxmlformats.org/package/2006/content-types"
_NS_MAIN    = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_REL_DOC = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

# Types de relations feuille → parties dessins/commentaires à restaurer,
# avec le tag XML à réinsérer dans la feuille (None pour comments : la
# relation suffit, le lien visuel passe par le vmlDrawing)
_REL_TYPES_DESSIN: Dict[str, Optional[str]] = {
    f"{_NS_REL_DOC}/drawing":     "drawing",
    f"{_NS_REL_DOC}/vmlDrawing":  "legacyDrawing",
    f"{_NS_REL_DOC}/comments":    None,
}

# Éléments de worksheet devant rester APRÈS <drawing>/<legacyDrawing>
# (ordre du schéma OOXML) — point d'insertion des tags restaurés
_APRES_DESSIN = ("<picture", "<oleObjects", "<controls", "<webPublishItems",
                 "<tableParts", "<extLst", "</worksheet>")


def _prefixe_sortie(client: str, date_cloture: str) -> str:
    """Construit le préfixe de sortie : '{annee}_12_{client}_'."""
    annee = pd.to_datetime(date_cloture, format="%d/%m/%Y").year
    return f"{annee}_12_{client}_"


def _normaliser(s: str) -> str:
    """
    Normalise pour la comparaison : sans accents, sans séparateurs, minuscules.
    Ex: 'X_Re_sultat' et 'X_Résultat exceptionnel' donnent tous deux 'xresultat…'.
    """
    import unicodedata
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")  # strip accents
    # Supprimer tous les séparateurs (espaces, underscores, tirets)
    for sep in (" ", "_", "-"):
        s = s.replace(sep, "")
    return s.lower()


def _detecter_cycle(nom_fichier: str, mapping_templates: Dict[str, str]) -> Optional[str]:
    """
    Retrouve le cycle d'un template à partir de son nom de fichier.

    La comparaison est normalisée (espaces → _, sans accents, minuscules)
    pour gérer les écarts entre les clés YAML et les noms de fichiers réels.
    """
    partie_norm = _normaliser(
        nom_fichier.replace(_PREFIXE_TEMPLATE, "").replace(".xlsx", "")
    )
    for cle, cycle in mapping_templates.items():
        if _normaliser(cle) in partie_norm:
            return cycle
    return None


def _feuilles_a_inserer(nom_fichier: str,
                        integration_templates: Dict[str, dict]) -> List[str]:
    """
    Retrouve les feuilles FM à insérer dans un template depuis la section
    integration_templates du YAML (clés = noms de fichiers réels).

    Les clés sont comparées normalisées, de la plus longue à la plus courte
    (évite qu'une clé courte capture un nom de fichier plus spécifique).
    """
    partie_norm = _normaliser(
        nom_fichier.replace(_PREFIXE_TEMPLATE, "").replace(".xlsx", "")
    )
    for cle in sorted(integration_templates, key=len, reverse=True):
        if _normaliser(cle) == partie_norm:
            return [str(f) for f in
                    (integration_templates[cle].get("feuilles") or [])]
    logger.warning(
        "Template '%s' : aucune entrée integration_templates ne correspond "
        "— aucune feuille FM insérée", nom_fichier,
    )
    return []


def _remplacer_placeholders(ws, nom_client: str, date_cloture: str) -> int:
    """
    Remplace les placeholders dans toutes les cellules de la feuille.
    Retourne le nombre de remplacements effectués.
    """
    nb = 0
    date_fmt = pd.to_datetime(date_cloture, format="%d/%m/%Y").strftime("%d/%m/%Y")

    substitutions = {
        "#NomClient":       nom_client,
        "#DateClôture":     date_fmt,
        "#Date":            date_fmt,
        "#date de clôture": date_fmt,
    }

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                valeur = cell.value
                for placeholder, remplacement in substitutions.items():
                    if placeholder in valeur:
                        valeur = valeur.replace(placeholder, remplacement)
                        nb += 1
                if valeur != cell.value:
                    cell.value = valeur

    return nb


# ---------------------------------------------------------------------------
# Insertion des feuilles FM (P1)
# ---------------------------------------------------------------------------

def _inserer_onglets_fm(wb_template, wb_fm, feuilles: List[str],
                        nom_template: str) -> int:
    """
    Insère les feuilles du FM listées comme nouveaux onglets du template,
    juste AVANT la feuille Synthèse, dans l'ordre de la liste.

    Paramètres
    ----------
    wb_template : Workbook
        Classeur template ouvert en écriture.
    wb_fm : Workbook
        Classeur FM source (chargé sans read_only pour disposer des styles).
    feuilles : list[str]
        Noms des onglets FM à copier (ex: ["AACE", "A0"]).
    nom_template : str
        Nom du fichier template (pour les logs).

    Retourne
    --------
    int
        Nombre de feuilles effectivement insérées.
    """
    nom_synth = next(
        (s for s in wb_template.sheetnames if "synth" in s.lower()), None,
    )
    if nom_synth is None:
        logger.warning(
            "Template '%s' : aucune feuille Synthèse trouvée — feuilles FM "
            "insérées en fin de classeur", nom_template,
        )

    nb_inserees = 0
    for nom_feuille in feuilles:
        if nom_feuille not in wb_fm.sheetnames:
            logger.warning(
                "Template '%s' : feuille FM '%s' absente du FM généré — "
                "ignorée", nom_template, nom_feuille,
            )
            continue
        index = (wb_template.sheetnames.index(nom_synth)
                 if nom_synth else None)
        ws_inseree = copy_worksheet(wb_template, wb_fm[nom_feuille],
                                    titre=nom_feuille, index=index)
        remove_gridlines(ws_inseree)  # Design A : aucun quadrillage
        nb_inserees += 1

    if nb_inserees:
        logger.info(
            "Template '%s' : %d feuille(s) FM insérée(s) avant '%s' : %s",
            nom_template, nb_inserees, nom_synth or "(fin)",
            ", ".join(feuilles),
        )
    return nb_inserees


# ---------------------------------------------------------------------------
# Restauration des dessins / médias / commentaires (manipulation ZIP/XML)
# ---------------------------------------------------------------------------

def _mapper_feuilles_zip(zf: zipfile.ZipFile) -> Dict[str, str]:
    """Retourne {nom_feuille: chemin_part} (ex: 'xl/worksheets/sheet3.xml')."""
    workbook = ElementTree.fromstring(zf.read("xl/workbook.xml"))
    rels = ElementTree.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    cibles = {
        rel.get("Id"): rel.get("Target")
        for rel in rels.findall(f"{{{_NS_REL_PKG}}}Relationship")
    }
    mapping: Dict[str, str] = {}
    for sheet in workbook.iter(f"{{{_NS_MAIN}}}sheet"):
        rid = sheet.get(f"{{{_NS_REL_DOC}}}id")
        cible = cibles.get(rid, "")
        if cible.startswith("/"):
            cible = cible.lstrip("/")
        else:
            cible = f"xl/{cible}"
        mapping[sheet.get("name")] = cible
    return mapping


def _rels_dessins_source(zf: zipfile.ZipFile,
                         part_feuille: str) -> List[Tuple[str, str]]:
    """Retourne [(type_rel, target)] des relations dessins/commentaires
    de la feuille (liste vide si la feuille n'a pas de fichier rels)."""
    nom = Path(part_feuille).name
    chemin_rels = f"xl/worksheets/_rels/{nom}.rels"
    if chemin_rels not in zf.namelist():
        return []
    rels = ElementTree.fromstring(zf.read(chemin_rels))
    return [
        (rel.get("Type"), rel.get("Target"))
        for rel in rels.findall(f"{{{_NS_REL_PKG}}}Relationship")
        if rel.get("Type") in _REL_TYPES_DESSIN
    ]


def _est_part_dessin(nom: str) -> bool:
    """True pour les parties à réinjecter depuis la source (drawings,
    médias, commentaires)."""
    return (nom.startswith("xl/drawings/") or nom.startswith("xl/media/")
            or re.fullmatch(r"xl/comments\d*\.xml", nom) is not None)


def _purger_refs_dessins(xml_feuille: str) -> str:
    """Supprime les tags <drawing/> et <legacyDrawing/> d'une feuille."""
    return re.sub(r"<(?:drawing|legacyDrawing)\b[^>]*/>", "", xml_feuille)


def _inserer_tags_dessins(xml_feuille: str, tags: List[str]) -> str:
    """Insère les tags dessins au bon endroit du XML de feuille (avant
    picture/oleObjects/…/extLst ou </worksheet>, ordre du schéma OOXML).

    Déclare le namespace r: sur l'élément racine <worksheet> si la feuille
    générée par openpyxl ne le porte pas (feuilles sans relation)."""
    if not tags:
        return xml_feuille
    debut_racine = xml_feuille.find("<worksheet")
    fin_racine = xml_feuille.find(">", debut_racine)
    if "xmlns:r=" not in xml_feuille[debut_racine:fin_racine]:
        xml_feuille = xml_feuille.replace(
            "<worksheet ", f'<worksheet xmlns:r="{_NS_REL_DOC}" ', 1,
        )
    bloc = "".join(tags)
    positions = [xml_feuille.find(b) for b in _APRES_DESSIN
                 if xml_feuille.find(b) != -1]
    point = min(positions)
    return xml_feuille[:point] + bloc + xml_feuille[point:]


def _fusionner_rels(xml_rels: Optional[bytes],
                    rels_source: List[Tuple[str, str]]) -> Tuple[bytes, Dict[str, str]]:
    """Reconstruit le fichier rels d'une feuille : retire les relations
    dessins/commentaires existantes, ajoute celles de la source avec des
    rIds libres. Retourne (xml, {type_rel: rid_attribué})."""
    ElementTree.register_namespace("", _NS_REL_PKG)
    if xml_rels is not None:
        racine = ElementTree.fromstring(xml_rels)
        for rel in list(racine):
            if rel.get("Type") in _REL_TYPES_DESSIN:
                racine.remove(rel)
    else:
        racine = ElementTree.Element(f"{{{_NS_REL_PKG}}}Relationships")

    existants = {rel.get("Id") for rel in racine}
    rids: Dict[str, str] = {}
    compteur = 1
    for type_rel, target in rels_source:
        while f"rId{compteur}" in existants:
            compteur += 1
        rid = f"rId{compteur}"
        existants.add(rid)
        ElementTree.SubElement(
            racine, f"{{{_NS_REL_PKG}}}Relationship",
            {"Id": rid, "Type": type_rel, "Target": target},
        )
        rids[type_rel] = rid
    xml = ElementTree.tostring(racine, encoding="UTF-8", xml_declaration=True)
    return xml, rids


def _fusionner_content_types(xml_sortie: bytes, xml_source: bytes) -> bytes:
    """Fusionne [Content_Types].xml : retire les déclarations dessins/
    commentaires de la sortie openpyxl, réinjecte celles de la source
    (Overrides des parties copiées + Defaults des extensions médias)."""
    ElementTree.register_namespace("", _NS_CT)
    sortie = ElementTree.fromstring(xml_sortie)
    source = ElementTree.fromstring(xml_source)

    for elem in list(sortie):
        part = elem.get("PartName", "")
        if part and _est_part_dessin(part.lstrip("/")):
            sortie.remove(elem)

    extensions = {
        elem.get("Extension")
        for elem in sortie.findall(f"{{{_NS_CT}}}Default")
    }
    for elem in source:
        tag = elem.tag.split("}")[-1]
        if tag == "Default" and elem.get("Extension") not in extensions:
            sortie.append(elem)
            extensions.add(elem.get("Extension"))
        elif tag == "Override":
            part = (elem.get("PartName") or "").lstrip("/")
            if _est_part_dessin(part):
                sortie.append(elem)
    return ElementTree.tostring(sortie, encoding="UTF-8", xml_declaration=True)


def _restaurer_dessins(bytes_source: bytes, bytes_sortie: bytes) -> bytes:
    """
    Réinjecte dans l'archive générée par openpyxl les images, dessins et
    commentaires du template source, perdus à l'ouverture+sauvegarde.

    Étapes (correspondance des feuilles par NOM, jamais par index — les
    feuilles FM insérées décalent la numérotation sheetN.xml) :
    1. retirer de la sortie toutes les parties xl/drawings/*, xl/media/*,
       xl/comments*.xml (openpyxl n'en conserve qu'une partie) ;
    2. copier verbatim ces parties depuis la source (avec leurs _rels) ;
    3. pour chaque feuille source portant des relations dessins : purger
       les tags/relations résiduels de la feuille de sortie, recréer les
       relations avec des rIds libres et réinsérer <drawing>/<legacyDrawing> ;
    4. fusionner [Content_Types].xml (Overrides + Defaults).
    """
    with zipfile.ZipFile(io.BytesIO(bytes_source)) as zf_src:
        feuilles_src = _mapper_feuilles_zip(zf_src)
        parts_dessins = {
            n: zf_src.read(n) for n in zf_src.namelist()
            if _est_part_dessin(n) or n.startswith("xl/drawings/_rels/")
        }
        rels_par_feuille = {
            nom: _rels_dessins_source(zf_src, part)
            for nom, part in feuilles_src.items()
        }
        ct_source = zf_src.read("[Content_Types].xml")

    if not parts_dessins:
        return bytes_sortie  # rien à restaurer

    with zipfile.ZipFile(io.BytesIO(bytes_sortie)) as zf_out:
        feuilles_out = _mapper_feuilles_zip(zf_out)
        entrees = {n: zf_out.read(n) for n in zf_out.namelist()}

    # 1. Purger les parties dessins partiellement conservées par openpyxl
    entrees = {n: data for n, data in entrees.items()
               if not (_est_part_dessin(n) or n.startswith("xl/drawings/_rels/"))}

    # 3. Reconstruire les références feuille par feuille
    for nom_feuille, rels_source in rels_par_feuille.items():
        part_out = feuilles_out.get(nom_feuille)
        if part_out is None or part_out not in entrees:
            if rels_source:
                logger.warning(
                    "Restauration dessins : feuille '%s' absente de la "
                    "sortie — dessins non rattachés", nom_feuille,
                )
            continue

        xml_feuille = entrees[part_out].decode("utf-8")
        xml_feuille = _purger_refs_dessins(xml_feuille)

        chemin_rels = f"xl/worksheets/_rels/{Path(part_out).name}.rels"
        if rels_source:
            xml_rels, rids = _fusionner_rels(entrees.get(chemin_rels),
                                             rels_source)
            entrees[chemin_rels] = xml_rels
            tags = [
                f'<{tag} r:id="{rids[type_rel]}"/>'
                for type_rel, tag in _REL_TYPES_DESSIN.items()
                if tag is not None and type_rel in rids
            ]
            xml_feuille = _inserer_tags_dessins(xml_feuille, tags)
        elif chemin_rels in entrees:
            xml_rels, _ = _fusionner_rels(entrees[chemin_rels], [])
            entrees[chemin_rels] = xml_rels

        entrees[part_out] = xml_feuille.encode("utf-8")

    # 2. Réinjecter les parties source + 4. content types
    entrees.update(parts_dessins)
    entrees["[Content_Types].xml"] = _fusionner_content_types(
        entrees["[Content_Types].xml"], ct_source,
    )

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for nom, data in entrees.items():
            zf.writestr(nom, data)

    logger.info(
        "Restauration dessins : %d partie(s) réinjectée(s) depuis la source",
        len(parts_dessins),
    )
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Traitement d'un template
# ---------------------------------------------------------------------------

def _traiter_template(
    chemin_source: Path,
    nom_client: str,
    date_cloture: str,
    prefixe_sortie: str,
    wb_fm=None,
    feuilles_fm: Optional[List[str]] = None,
) -> Optional[tuple]:
    """
    Charge, modifie et retourne (nom_fichier_sortie, bytes_contenu) pour un template.
    Retourne None si le fichier ne peut pas être traité.

    Paramètres supplémentaires
    --------------------------
    wb_fm : Workbook ou None
        Classeur FM source (déjà chargé) — pour l'insertion d'onglets.
    feuilles_fm : list[str] ou None
        Noms des onglets FM à insérer avant la feuille Synthèse.
    """
    wb = openpyxl.load_workbook(chemin_source)

    total_remplacements = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        remove_gridlines(ws)
        total_remplacements += _remplacer_placeholders(ws, nom_client, date_cloture)

    # Insertion des feuilles maîtresses comme onglets (P1)
    if wb_fm is not None and feuilles_fm:
        _inserer_onglets_fm(wb, wb_fm, feuilles_fm, chemin_source.name)

    nom_sortie = chemin_source.name.replace(_PREFIXE_TEMPLATE, prefixe_sortie)

    # Sérialiser en mémoire
    buffer = io.BytesIO()
    wb.save(buffer)
    contenu = buffer.getvalue()

    # Restaurer les images/dessins/commentaires perdus par openpyxl
    contenu = _restaurer_dessins(chemin_source.read_bytes(), contenu)

    logger.info(
        "Template '%s' → '%s' (%d feuilles, %d remplacements)",
        chemin_source.name, nom_sortie, len(wb.sheetnames), total_remplacements,
    )
    return nom_sortie, contenu


def write(
    templates_dir: Union[str, Path],
    nom_client: str,
    date_cloture: str,
    output_path: Union[str, Path],
    mapping_templates: Optional[Dict[str, str]] = None,
    fm_path: Optional[Union[str, Path]] = None,
    integration_templates: Optional[Dict[str, dict]] = None,
) -> Path:
    """
    Traite tous les templates et génère un ZIP de feuilles de travail.

    Paramètres
    ----------
    templates_dir : str ou Path
        Dossier contenant les fichiers 20XX_XX_YYY_*.xlsx.
    nom_client : str
        Nom du client (ex: "GILAC").
    date_cloture : str
        Date de clôture au format 'JJ/MM/AAAA'.
    output_path : str ou Path
        Dossier de sortie où sera créé le ZIP.
    mapping_templates : dict, optionnel
        {cle_fichier: cycle} depuis mapping_pcg.yaml (section 'templates').
        Si None, tous les .xlsx du dossier sont traités sans filtrage par cycle.
    fm_path : str ou Path, optionnel
        Chemin du FM généré (fm_writer.write). Si fourni avec
        integration_templates, les feuilles maîtresses sont insérées comme
        onglets dans chaque template, avant la feuille Synthèse.
    integration_templates : dict, optionnel
        Section 'integration_templates' du mapping_pcg.yaml :
        {nom_fichier: {feuilles: [...], position: avant_synthese}}.

    Retourne
    --------
    Path
        Chemin vers le fichier ZIP généré.
    """
    templates_dir = Path(templates_dir)
    output_dir    = Path(output_path)
    output_dir.mkdir(parents=True, exist_ok=True)

    if not templates_dir.exists():
        raise FileNotFoundError(f"Dossier templates introuvable : {templates_dir}")

    templates = sorted(templates_dir.glob("*.xlsx"))
    if not templates:
        raise ValueError(f"Aucun template .xlsx trouvé dans {templates_dir}")

    annee   = pd.to_datetime(date_cloture, format="%d/%m/%Y").year
    prefixe = _prefixe_sortie(nom_client, date_cloture)
    nom_zip = output_dir / f"FT_{nom_client}_{annee}.zip"

    # Charger le FM une seule fois (sans read_only : les styles des onglets
    # copiés doivent être accessibles)
    wb_fm = None
    if fm_path is not None and integration_templates:
        fm_path = Path(fm_path)
        if fm_path.exists():
            wb_fm = openpyxl.load_workbook(fm_path)
            logger.info("FM chargé pour insertion : %s (%d onglets)",
                        fm_path.name, len(wb_fm.sheetnames))
        else:
            logger.warning(
                "FM introuvable : %s — insertion des feuilles FM ignorée",
                fm_path,
            )
    elif fm_path is not None:
        logger.warning(
            "fm_path fourni sans integration_templates — insertion des "
            "feuilles FM ignorée (vérifier mapping_pcg.yaml)."
        )

    nb_traites = 0
    nb_ignores = 0

    with zipfile.ZipFile(nom_zip, "w", zipfile.ZIP_DEFLATED) as zf:
        for chemin in templates:
            cycle = None

            # Détection du cycle (filtrage des fichiers hors mapping)
            if mapping_templates:
                cycle = _detecter_cycle(chemin.name, mapping_templates)
                if cycle is None:
                    logger.warning(
                        "Template '%s' ignoré : aucun cycle détecté dans le mapping",
                        chemin.name,
                    )
                    nb_ignores += 1
                    continue

            feuilles_fm = (
                _feuilles_a_inserer(chemin.name, integration_templates)
                if wb_fm is not None else None
            )

            resultat = _traiter_template(
                chemin_source=chemin,
                nom_client=nom_client,
                date_cloture=date_cloture,
                prefixe_sortie=prefixe,
                wb_fm=wb_fm,
                feuilles_fm=feuilles_fm,
            )
            if resultat is None:
                nb_ignores += 1
                continue

            nom_sortie, contenu = resultat
            zf.writestr(nom_sortie, contenu)
            nb_traites += 1

    logger.info(
        "ZIP généré : %s (%d templates traités, %d ignorés)",
        nom_zip, nb_traites, nb_ignores,
    )
    return nom_zip
