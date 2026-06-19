"""
Copie d'un onglet Excel entre deux classeurs openpyxl.

openpyxl ne copie PAS d'onglet entre workbooks (Workbook.copy_worksheet est
intra-classeur uniquement) : la feuille est recréée cellule par cellule en
copiant valeurs, styles, fusions et dimensions.

Détails techniques :
- Styles : copiés attribut par attribut avec copy.copy() (font, border,
  fill, number_format, alignment, protection) — ne JAMAIS recopier
  cell._style (index vers les tables de styles du classeur source).
- Insertion positionnée : wb.create_sheet(title, index=...) (préféré à
  move_sheet, plus fiable pour insérer avant une feuille donnée).
"""

import copy
import logging
from typing import Optional

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def copy_worksheet(
    wb_cible: Workbook,
    ws_source: Worksheet,
    titre: Optional[str] = None,
    index: Optional[int] = None,
) -> Worksheet:
    """Copie une feuille d'un classeur source vers un classeur cible.

    Recrée la feuille cellule par cellule : valeurs, styles (font, border,
    fill, number_format, alignment, protection), plages fusionnées,
    largeurs de colonnes, hauteurs de lignes, quadrillage et volets figés.

    Paramètres
    ----------
    wb_cible : Workbook
        Classeur openpyxl de destination (mode normal, pas read_only).
    ws_source : Worksheet
        Feuille source (d'un AUTRE classeur, chargé sans read_only pour
        disposer des styles).
    titre : str, optionnel
        Nom de la feuille créée dans la cible (défaut : titre de la source).
    index : int, optionnel
        Position d'insertion dans wb_cible (défaut : à la fin).

    Retourne
    --------
    Worksheet
        La feuille créée dans le classeur cible.

    Lève
    ----
    ValueError
        Si une feuille du même nom existe déjà dans le classeur cible.
    """
    titre = titre or ws_source.title
    if titre in wb_cible.sheetnames:
        raise ValueError(
            f"copy_worksheet : la feuille '{titre}' existe déjà dans le "
            f"classeur cible — renommer la feuille source ou la cible."
        )

    ws_cible = wb_cible.create_sheet(title=titre, index=index)

    # --- Cellules : valeurs + styles attribut par attribut ---
    for row in ws_source.iter_rows():
        for cell in row:
            nouvelle = ws_cible.cell(row=cell.row, column=cell.column,
                                     value=cell.value)
            if cell.has_style:
                nouvelle.font          = copy.copy(cell.font)
                nouvelle.border        = copy.copy(cell.border)
                nouvelle.fill          = copy.copy(cell.fill)
                nouvelle.number_format = cell.number_format
                nouvelle.alignment     = copy.copy(cell.alignment)
                nouvelle.protection    = copy.copy(cell.protection)

    # --- Plages fusionnées ---
    for plage in ws_source.merged_cells.ranges:
        ws_cible.merge_cells(str(plage))

    # --- Dimensions ---
    for lettre, dim in ws_source.column_dimensions.items():
        ws_cible.column_dimensions[lettre].width  = dim.width
        ws_cible.column_dimensions[lettre].hidden = dim.hidden
    for numero, dim in ws_source.row_dimensions.items():
        ws_cible.row_dimensions[numero].height = dim.height
        ws_cible.row_dimensions[numero].hidden = dim.hidden

    # --- Vue : quadrillage et volets figés ---
    ws_cible.sheet_view.showGridLines = ws_source.sheet_view.showGridLines
    ws_cible.freeze_panes = ws_source.freeze_panes

    logger.debug(
        "copy_worksheet : '%s' copiée (%d lignes × %d colonnes, %d fusions)",
        titre, ws_source.max_row, ws_source.max_column,
        len(ws_source.merged_cells.ranges),
    )
    return ws_cible
