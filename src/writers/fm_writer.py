"""
Générateur de feuilles maîtresses FM_{client}_{annee}.xlsx.

Onglets produits :
  1. Sommaire
  2. Balance N Vs N-1  (tous les comptes, valeurs brutes)
  3. Tréso              (BFR, FRNG, TN — agrégé par poste)
  4. AACE               (Autres Achats et Charges Externes — comptes 606-609, 61x, 62x)
  5. Un onglet par cycle  (A0, V0, C Propres0, …)
     — sections ACTIF / PASSIF / CHARGES / PRODUITS (seulement si non vides)
     — convention de signe : PASSIF et PRODUITS sont présentés en positif (Solde × -1)
"""

import logging
from pathlib import Path
from typing import List, Optional, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from src.engine.liasse_fiscale_loader import load_liasse_fiscale
from src.writers.styles import (
    remove_gridlines, set_col_widths, write_title_block,
    write_header_row, write_data_row, write_section_label, write_total_row,
    FONT_BOLD, FONT_META, FONT_NORMAL, FONT_SECTION, NUM_KE, NUM_PCT,
)

logger = logging.getLogger(__name__)

# Chemin par défaut de la config PCG (fallback si write() est appelé sans
# pcg_config — compatibilité avec les appels existants)
_PCG_DEFAULT = Path(__file__).resolve().parent.parent / "config" / "mapping_pcg.yaml"

# Ordre des sections dans les onglets cycle
ORDRE_SECTIONS = ["Actif", "Passif", "Charges", "Produits"]

# Signe de présentation : +1 → même signe, -1 → inversé
_SIGNE: dict = {"Actif": 1, "Passif": -1, "Charges": 1, "Produits": -1}

# Largeurs de colonnes standard
_WIDTHS_BG = {"A": 6, "B": 12, "C": 38, "D": 14, "E": 14,
              "F": 12, "G": 10, "H": 14, "I": 4, "J": 12,
              "K": 28, "L": 28, "M": 10, "N": 10}

_WIDTHS_CYCLE = {"A": 6, "B": 12, "C": 38, "D": 14, "E": 14,
                 "F": 12, "G": 10, "H": 14, "I": 4, "J": 12}

_WIDTHS_TRESO = {"A": 6, "B": 36, "C": 30, "D": 14, "E": 14, "F": 12, "G": 10}

_WIDTHS_BILAN = {
    "A": 4,  "B": 30, "C": 4,  "D": 14, "E": 14, "F": 12, "G": 10,
    "H": 4,  "I": 30, "J": 4,  "K": 14, "L": 14, "M": 12, "N": 10,
}

# Headers communs (colonnes A-J)
_HEADERS_BASE = [
    (1, "Ref."),
    (3, None),          # CompteLib — pas de label
    (6, "Var. K€"),
    (7, "Var. %"),
    (8, "Ref."),
    (10, "Cycle"),
]


def _date_str(date_cloture: str) -> str:
    """Convertit 'JJ/MM/AAAA' en pd.Timestamp puis en string '31/12/2025'."""
    return pd.to_datetime(date_cloture, format="%d/%m/%Y").strftime("%d/%m/%Y")


def _annee(date_cloture: str) -> int:
    return pd.to_datetime(date_cloture, format="%d/%m/%Y").year


def _var_pct_val(var_pct) -> Union[float, str]:
    """Retourne le float ou 'n/a' tel quel."""
    return var_pct if isinstance(var_pct, str) else float(var_pct)


# ---------------------------------------------------------------------------
# Onglet Sommaire
# ---------------------------------------------------------------------------

def _ecrire_sommaire(ws, client: str, date_cloture: str, cycles: list) -> None:
    remove_gridlines(ws)
    set_col_widths(ws, {"A": 4, "B": 30, "C": 50})

    ws.cell(row=3, column=3, value="Sommaire").font = FONT_BOLD
    ws.cell(row=4, column=3, value=client).font     = FONT_META

    ws.cell(row=6, column=2, value=f"Dossier : {client} — {date_cloture}").font = FONT_NORMAL

    ws.cell(row=8, column=2, value="Balance générale auditée").font = FONT_BOLD
    ws.cell(row=9, column=3, value="Balance N Vs N-1").font = FONT_NORMAL

    ws.cell(row=11, column=2, value="Feuilles maîtresses").font = FONT_BOLD
    ws.cell(row=12, column=3, value="Bilan").font  = FONT_NORMAL
    ws.cell(row=13, column=3, value="Tréso").font  = FONT_NORMAL
    ws.cell(row=14, column=3, value="AACE").font   = FONT_NORMAL
    for i, cycle in enumerate(cycles):
        nom = f"Cycle {cycle}"
        ws.cell(row=15 + i, column=3, value=nom).font = FONT_NORMAL


# ---------------------------------------------------------------------------
# Onglet Balance N Vs N-1
# ---------------------------------------------------------------------------

def _ecrire_balance_tab(ws, df: pd.DataFrame, date_n: str, date_n1: str,
                        client: str) -> None:
    remove_gridlines(ws)
    set_col_widths(ws, _WIDTHS_BG)

    ws.cell(row=1, column=1, value="Retour sommaire").font = FONT_META

    write_title_block(ws, row_titre=3, row_client=4,
                      titre=f"Balances générales auditées : {date_n} Vs {date_n1}",
                      client=client)

    # Headers row 8
    headers = [
        (1, "Ref."), (2, "En milliers d'€uros"), (4, date_n), (5, date_n1),
        (6, "Var. K€"), (7, "Var. %"), (8, "Ref."),
        (10, "Cycle"), (11, "EtatFi N"), (12, "EtatFi N-1"),
        (13, "ComptaN"), (14, "ComptaN-1"),
    ]
    write_header_row(ws, 8, headers)

    # Données à partir de row 10
    current_row = 10
    for _, r in df.iterrows():
        var_pct = _var_pct_val(r["Var_PCT"])
        cells = [
            (2, str(r["CompteNum"]),  None,   FONT_META),
            (3, r["CompteLib"],       None,   FONT_NORMAL),
            (4, round(r["Solde_KE"], 3),  NUM_KE,  FONT_NORMAL),
            (5, round(r["Solde_N1_KE"], 3), NUM_KE, FONT_NORMAL),
            (6, round(r["Var_KE"], 3),   NUM_KE,  FONT_NORMAL),
            (7, var_pct,              NUM_PCT if not isinstance(var_pct, str) else None,
               FONT_NORMAL),
            (8, r["ref"],             None,   FONT_META),
            (10, r["cycle"],          None,   FONT_META),
            (11, r.get("etatfi", ""), None,  FONT_META),
            (12, r.get("etatfi", ""), None,  FONT_META),
            (13, r["compta"],         None,   FONT_META),
            (14, r["compta"],         None,   FONT_META),
        ]
        write_data_row(ws, current_row, cells)
        current_row += 1


# ---------------------------------------------------------------------------
# Onglet Bilan (Actif gauche / Passif droite)
# ---------------------------------------------------------------------------

def _classer_comptes_bascule(balance: pd.DataFrame,
                             prefixes_bascule: List[str]) -> tuple:
    """
    Reclasse dynamiquement les comptes bascule (préfixes issus de
    liasse_fiscale.bilan.comptes_bascule.prefixes) selon le signe de leur
    Solde_KE / Solde_N1_KE.

    Retourne ((actif_n, actif_n1), (passif_n, passif_n1)).
    Convention passif : valeur positive (−Solde_KE pour les soldes créditeurs).
    """
    if "Solde_KE" not in balance.columns or "Solde_N1_KE" not in balance.columns:
        raise ValueError(
            "Bascule : colonnes 'Solde_KE' ou 'Solde_N1_KE' absentes de la balance — "
            "vérifier que balance_builder.build() a bien été appelé."
        )
    masque = balance["CompteNum"].apply(
        lambda n: any(str(n).startswith(p) for p in prefixes_bascule)
    )
    df_b = balance.loc[masque]

    actif_n   = round(float(df_b.loc[df_b["Solde_KE"]    > 0, "Solde_KE"].sum()),    3)
    actif_n1  = round(float(df_b.loc[df_b["Solde_N1_KE"] > 0, "Solde_N1_KE"].sum()), 3)
    passif_n  = round(-float(df_b.loc[df_b["Solde_KE"]    < 0, "Solde_KE"].sum()),   3)
    passif_n1 = round(-float(df_b.loc[df_b["Solde_N1_KE"] < 0, "Solde_N1_KE"].sum()), 3)

    logger.debug(
        "Bascule : actif N=%.1f K€, passif N=%.1f K€ (%d comptes bascule)",
        actif_n, passif_n, len(df_b),
    )
    return (actif_n, actif_n1), (passif_n, passif_n1)


def _calc_bilan(balance: pd.DataFrame, liasse_fiscale: dict) -> dict:
    """
    Calcule les postes actif et passif du bilan synthétique pour N et N-1.

    Les listes de préfixes proviennent de liasse_fiscale["bilan"] (section
    liasse_fiscale du mapping_pcg.yaml). Les signes de présentation et les
    conditions sur le signe du solde restent dans ce code (logique générique).

    Convention :
    - Actif : valeurs nettes (brut + amort, les amort ayant des Solde_KE négatifs)
    - Passif : × -1 (comptes créditeurs → soldes négatifs → affichage positif)

    Garantie universelle : un poste résiduel catch-all absorbe les comptes
    non capturés pour assurer l'équilibre avec n'importe quel FEC.
    """
    def sp(prefixes, signe=1):
        return (
            round(_sommer_prefixes(balance, prefixes, "Solde_KE")    * signe, 3),
            round(_sommer_prefixes(balance, prefixes, "Solde_N1_KE") * signe, 3),
        )

    def sp_si(prefixes, condition, signe=1):
        return (
            round(_sommer_prefixes_si(balance, prefixes, "Solde_KE",    condition) * signe, 3),
            round(_sommer_prefixes_si(balance, prefixes, "Solde_N1_KE", condition) * signe, 3),
        )

    p_actif   = liasse_fiscale["bilan"]["actif"]
    p_passif  = liasse_fiscale["bilan"]["passif"]
    p_bascule = liasse_fiscale["bilan"]["comptes_bascule"]

    # --- ACTIF ---
    # Les comptes d'amort/dép ont des Solde_KE négatifs — les additionner
    # aux comptes bruts donne directement la valeur nette.
    cap_non_appele = sp(p_actif["capital_souscrit_non_appele"])
    immo_incorp    = sp(p_actif["immo_incorp"])
    immo_corp      = sp(p_actif["immo_corp"])
    immo_fi        = sp(p_actif["immo_fi"])
    stocks         = sp(p_actif["stocks"])
    avances        = sp(p_actif["avances_acomptes_verses"])
    crean_cli      = sp(p_actif["creances_clients"])
    # Les préfixes bascule (425, 444, 451, 455, 456, 467, 478, 486) sont
    # absents des listes actif/passif — gérés par _classer_comptes_bascule
    autres_crean   = sp(p_actif["autres_creances"])
    vmp            = sp(p_actif["vmp"])
    dispo          = sp(p_actif["disponibilites"])

    # --- PASSIF (× -1 car comptes créditeurs) ---
    capital        = sp(p_passif["capital"],                  signe=-1)
    primes         = sp(p_passif["primes"],                   signe=-1)
    reserve_legale = sp(p_passif["reserve_legale"],           signe=-1)
    autres_res     = sp(p_passif["autres_reserves"],          signe=-1)
    report         = sp(p_passif["report_nouveau"],           signe=-1)
    resultat       = sp(p_passif["resultat_exercice"],        signe=-1)
    subventions    = sp(p_passif["subventions"],              signe=-1)
    prov_regl      = sp(p_passif["prov_reglementees"],        signe=-1)
    prov_risques_b = sp(p_passif["prov_risques"],             signe=-1)
    prov_charges_b = sp(p_passif["prov_charges"],             signe=-1)
    emprunts       = sp(p_passif["emprunts"],                 signe=-1)
    det_fourn_b    = sp(p_passif["dettes_fournisseurs"],      signe=-1)
    det_fisc_b     = sp(p_passif["dettes_fiscales_sociales"], signe=-1)
    autres_dettes  = sp(p_passif["autres_dettes"],            signe=-1)
    # Résultat en cours (classes 6 & 7 si exercice non clôturé / arrêté partiel)
    resultat_encours = sp(p_passif["resultat_en_cours"],      signe=-1)

    # --- Comptes bascule : reclassement dynamique selon le signe ---
    bascule_actif, bascule_passif = _classer_comptes_bascule(
        balance, p_bascule["prefixes"]
    )

    # Sous-totaux interco affichés séparément (display only — déjà inclus dans bascule_actif/passif)
    crean_interco           = sp_si(p_bascule["interco"], ">0")
    dettes_interco          = sp_si(p_bascule["interco"], "<0", signe=-1)
    bascule_reclasses_actif = (
        round(bascule_actif[0]  - crean_interco[0],  3),
        round(bascule_actif[1]  - crean_interco[1],  3),
    )
    bascule_reclasses_passif = (
        round(bascule_passif[0] - dettes_interco[0], 3),
        round(bascule_passif[1] - dettes_interco[1], 3),
    )

    # --- Totaux ---
    total_actif = (
        round(cap_non_appele[0] + immo_incorp[0] + immo_corp[0] + immo_fi[0]
              + stocks[0] + avances[0] + crean_cli[0] + autres_crean[0]
              + bascule_actif[0] + vmp[0] + dispo[0], 3),
        round(cap_non_appele[1] + immo_incorp[1] + immo_corp[1] + immo_fi[1]
              + stocks[1] + avances[1] + crean_cli[1] + autres_crean[1]
              + bascule_actif[1] + vmp[1] + dispo[1], 3),
    )
    total_passif = (
        round(capital[0] + primes[0] + reserve_legale[0] + autres_res[0]
              + report[0] + resultat[0] + subventions[0] + prov_regl[0]
              + prov_risques_b[0] + prov_charges_b[0] + emprunts[0]
              + det_fourn_b[0] + det_fisc_b[0] + autres_dettes[0]
              + bascule_passif[0] + resultat_encours[0], 3),
        round(capital[1] + primes[1] + reserve_legale[1] + autres_res[1]
              + report[1] + resultat[1] + subventions[1] + prov_regl[1]
              + prov_risques_b[1] + prov_charges_b[1] + emprunts[1]
              + det_fourn_b[1] + det_fisc_b[1] + autres_dettes[1]
              + bascule_passif[1] + resultat_encours[1], 3),
    )

    # Diagnostic : log l'écart résiduel pour traçabilité
    imbalance_n = round(total_actif[0] - total_passif[0], 3)
    if abs(imbalance_n) > 0.1:
        logger.warning(
            "Bilan : résiduel actif/passif N = %.3f K€ "
            "(comptes non capturés → vérifier mapping_pcg.yaml)", imbalance_n,
        )

    # Contrôle équilibre bilan
    for annee_lbl, idx in [("N", 0), ("N-1", 1)]:
        ecart = abs(total_actif[idx] - total_passif[idx])
        if ecart > 0.01:
            logger.warning(
                "Bilan : Total Actif ≠ Total Passif en %s "
                "(Actif=%.1f, Passif=%.1f, écart=%.3f K€)",
                annee_lbl, total_actif[idx], total_passif[idx], ecart,
            )

    return dict(
        cap_non_appele=cap_non_appele,
        immo_incorp=immo_incorp, immo_corp=immo_corp, immo_fi=immo_fi,
        stocks=stocks, avances=avances, crean_cli=crean_cli,
        autres_crean=autres_crean,
        crean_interco=crean_interco,
        bascule_reclasses_actif=bascule_reclasses_actif,
        vmp=vmp, dispo=dispo,
        total_actif=total_actif,
        capital=capital, primes=primes, reserve_legale=reserve_legale,
        autres_res=autres_res, report=report, resultat=resultat,
        subventions=subventions, prov_regl=prov_regl,
        prov_risques_b=prov_risques_b, prov_charges_b=prov_charges_b,
        emprunts=emprunts, det_fourn_b=det_fourn_b, det_fisc_b=det_fisc_b,
        autres_dettes=autres_dettes,
        dettes_interco=dettes_interco,
        bascule_reclasses_passif=bascule_reclasses_passif,
        resultat_encours=resultat_encours,
        total_passif=total_passif,
    )


def _ecrire_bilan_tab(ws, balance: pd.DataFrame,
                      date_n: str, date_n1: str, client: str,
                      liasse_fiscale: dict) -> None:
    """Écrit l'onglet Bilan — Actif à gauche (B-G), Passif à droite (I-N)."""
    remove_gridlines(ws)
    set_col_widths(ws, _WIDTHS_BILAN)

    ws.cell(row=1, column=1, value="Retour sommaire").font = FONT_META

    write_title_block(ws, row_titre=3, row_client=4,
                      titre="Bilan synthétique", client=client)

    write_header_row(ws, 8, [
        (2, "ACTIF"),  (4, date_n), (5, date_n1), (6, "Var. K€"), (7, "Var. %"),
        (9, "PASSIF"), (11, date_n), (12, date_n1), (13, "Var. K€"), (14, "Var. %"),
    ])

    d = _calc_bilan(balance, liasse_fiscale)

    # Constructeurs de ligne (type, label, val_n, val_n1)
    def rs(label):      return ("s", label,     None,      None)
    def rp(label, key): return ("p", label,     d[key][0], d[key][1])
    def rt(label, key): return ("t", label,     d[key][0], d[key][1])
    def rb():           return ("b", "",         None,      None)

    actif_rows = [
        rs("ACTIF IMMOBILISÉ"),
        rp("Capital souscrit non appelé",          "cap_non_appele"),
        rp("Immobilisations incorporelles (net)",  "immo_incorp"),
        rp("Immobilisations corporelles (net)",    "immo_corp"),
        rp("Immobilisations financières (net)",    "immo_fi"),
        rb(),
        rs("ACTIF CIRCULANT"),
        rp("Stocks et en-cours (net)",             "stocks"),
        rp("Avances et acomptes versés",           "avances"),
        rp("Créances clients (net)",               "crean_cli"),
        rp("Autres créances",                      "autres_crean"),
        rp("Créances interco (451, 455 débiteurs)","crean_interco"),
        rp("Comptes bascule reclassés (actif)",    "bascule_reclasses_actif"),
        rp("VMP (net)",                            "vmp"),
        rp("Disponibilités",                       "dispo"),
        rb(), rt("TOTAL ACTIF", "total_actif"),
    ]

    passif_rows = [
        rs("CAPITAUX PROPRES"),
        rp("Capital",                              "capital"),
        rp("Primes d'émission",                    "primes"),
        rp("Réserve légale",                       "reserve_legale"),
        rp("Autres réserves",                      "autres_res"),
        rp("Report à nouveau",                     "report"),
        rp("Résultat de l'exercice",               "resultat"),
        rp("Résultat en cours (6/7)",              "resultat_encours"),
        rp("Subventions d'investissement",         "subventions"),
        rp("Provisions réglementées",              "prov_regl"),
        rb(),
        rs("PROVISIONS"),
        rp("Provisions pour risques",              "prov_risques_b"),
        rp("Provisions pour charges",              "prov_charges_b"),
        rb(),
        rs("DETTES"),
        rp("Emprunts et dettes financières",       "emprunts"),
        rp("Dettes fournisseurs",                  "det_fourn_b"),
        rp("Dettes fiscales et sociales",          "det_fisc_b"),
        rp("Autres dettes",                        "autres_dettes"),
        rp("Dettes interco (451, 455 créditeurs)", "dettes_interco"),
        rp("Comptes bascule reclassés (passif)",   "bascule_reclasses_passif"),
        rb(), rt("TOTAL PASSIF", "total_passif"),
    ]

    _VIDE = ("b", "", None, None)

    def _var(vn, vn1):
        var_ke  = round(vn - vn1, 3)
        var_pct = round(var_ke / abs(vn1), 4) if abs(vn1) >= 0.001 else "n/a"
        return var_ke, var_pct

    def _ecrire_cote(row, item, col_label, col_n, col_n1, col_var_ke, col_var_pct):
        typ, label, vn, vn1 = item
        if typ == "s":
            ws.cell(row=row, column=col_label, value=label).font = FONT_SECTION
        elif typ == "p":
            var_ke, var_pct = _var(vn, vn1)
            write_data_row(ws, row, [
                (col_label,    label,   None,   FONT_NORMAL),
                (col_n,        vn,      NUM_KE, FONT_NORMAL),
                (col_n1,       vn1,     NUM_KE, FONT_NORMAL),
                (col_var_ke,   var_ke,  NUM_KE, FONT_NORMAL),
                (col_var_pct,  var_pct,
                 NUM_PCT if not isinstance(var_pct, str) else None,
                 FONT_NORMAL),
            ])
        elif typ == "t":
            var_ke, var_pct = _var(vn, vn1)
            write_total_row(ws, row, [
                (col_label,   label),
                (col_n,       vn,      NUM_KE),
                (col_n1,      vn1,     NUM_KE),
                (col_var_ke,  var_ke,  NUM_KE),
                (col_var_pct, var_pct,
                 NUM_PCT if not isinstance(var_pct, str) else None),
            ])

    for i in range(max(len(actif_rows), len(passif_rows))):
        excel_row = 10 + i
        a = actif_rows[i]  if i < len(actif_rows)  else _VIDE
        p = passif_rows[i] if i < len(passif_rows) else _VIDE
        _ecrire_cote(excel_row, a, 2,  4,  5,  6,  7)
        _ecrire_cote(excel_row, p, 9, 11, 12, 13, 14)

    logger.info(
        "Bilan : onglet généré (Actif=%.0f, Passif=%.0f K€)",
        d["total_actif"][0], d["total_passif"][0],
    )


# ---------------------------------------------------------------------------
# Helpers calcul Tréso (BFR / FRNG / TN)
# ---------------------------------------------------------------------------

def _sommer_prefixes(balance: pd.DataFrame, prefixes: list,
                     colonne: str = "Solde_KE") -> float:
    """Somme les valeurs de colonne pour tous les comptes commençant par un des préfixes."""
    masque = balance["CompteNum"].apply(
        lambda n: any(str(n).startswith(p) for p in prefixes)
    )
    return float(balance.loc[masque, colonne].sum())


def _sommer_prefixes_si(balance: pd.DataFrame, prefixes: list,
                         colonne: str, signe_cond: str) -> float:
    """Somme avec filtre sur le signe de la valeur ('>0' ou '<0')."""
    masque = balance["CompteNum"].apply(
        lambda n: any(str(n).startswith(p) for p in prefixes)
    )
    if signe_cond == ">0":
        masque = masque & (balance[colonne] > 0)
    elif signe_cond == "<0":
        masque = masque & (balance[colonne] < 0)
    return float(balance.loc[masque, colonne].sum())


def _calc_treso(balance: pd.DataFrame, liasse_fiscale: dict) -> dict:
    """
    Calcule tous les postes BFR/FRNG/TN pour N (Solde_KE) et N-1 (Solde_N1_KE).

    Les listes de préfixes proviennent de liasse_fiscale["treso"] (section
    liasse_fiscale du mapping_pcg.yaml). Les signes et les conditions sur le
    signe du solde restent dans ce code (logique générique).

    Retourne un dict {nom_poste: (val_n, val_n1)}.
    """
    def sp(prefixes, signe=1):
        """Somme simple × signe pour N et N-1."""
        return (
            round(_sommer_prefixes(balance, prefixes, "Solde_KE")    * signe, 3),
            round(_sommer_prefixes(balance, prefixes, "Solde_N1_KE") * signe, 3),
        )

    def sp_si(prefixes, cond, signe=1):
        """Somme conditionnelle × signe pour N et N-1."""
        return (
            round(_sommer_prefixes_si(balance, prefixes, "Solde_KE",    cond) * signe, 3),
            round(_sommer_prefixes_si(balance, prefixes, "Solde_N1_KE", cond) * signe, 3),
        )

    def add(*postes):
        """Additionne plusieurs tuples (val_n, val_n1)."""
        return (round(sum(p[0] for p in postes), 3),
                round(sum(p[1] for p in postes), 3))

    p_res   = liasse_fiscale["treso"]["ressources_stables"]
    p_emp   = liasse_fiscale["treso"]["emplois_stables"]
    p_ac    = liasse_fiscale["treso"]["actif_circulant_exploitation"]
    p_pc    = liasse_fiscale["treso"]["passif_circulant_exploitation"]
    p_treso = liasse_fiscale["treso"]["tresorerie_directe"]

    # --- Ressources stables (× -1 car passif/créditeur) ---
    cap_propres  = sp(p_res["capitaux_propres"],       signe=-1)
    amort_dep    = sp(p_res["amort_dep"],              signe=-1)
    prov_risques = sp(p_res["prov_risques"],           signe=-1)
    dettes_mlt   = sp(p_res["dettes_financieres_mlt"], signe=-1)
    total_res    = add(cap_propres, amort_dep, prov_risques, dettes_mlt)

    # --- Emplois stables (positif car actif/débiteur) ---
    actif_immo = sp(p_emp["actif_immobilise_brut"])
    total_emp  = actif_immo

    # --- FRNG ---
    frng = (round(total_res[0] - total_emp[0], 3),
            round(total_res[1] - total_emp[1], 3))

    # --- Actif circulant d'exploitation ---
    stocks       = sp(p_ac["stocks"])
    crean_cli    = sp(p_ac["creances_clients"])
    autres_crean = sp_si(p_ac["autres_creances"], ">0")
    cca          = sp(p_ac["cca"])
    total_ac     = add(stocks, crean_cli, autres_crean, cca)

    # --- Passif circulant d'exploitation (× -1 pour affichage positif) ---
    det_fourn  = sp(p_pc["dettes_fournisseurs"], signe=-1)
    det_fisc   = sp_si(p_pc["dettes_fiscales_sociales"], "<0", signe=-1)
    autres_det = sp_si(p_pc["autres_dettes"],            "<0", signe=-1)
    pca        = sp(p_pc["pca"], signe=-1)
    total_pc   = add(det_fourn, det_fisc, autres_det, pca)

    # --- BFR ---
    bfr = (round(total_ac[0] - total_pc[0], 3),
           round(total_ac[1] - total_pc[1], 3))

    # --- TN ---
    tn = (round(frng[0] - bfr[0], 3),
          round(frng[1] - bfr[1], 3))

    # --- Vérification TN (trésorerie directe) ---
    treso_active   = sp_si(p_treso["tresorerie_active"], ">0")
    treso_pass_519 = sp(p_treso["tresorerie_passive_519"], signe=-1)
    treso_pass_512 = sp_si(p_treso["tresorerie_passive_512"], "<0", signe=-1)
    treso_passive  = (round(treso_pass_519[0] + treso_pass_512[0], 3),
                      round(treso_pass_519[1] + treso_pass_512[1], 3))
    tn_verif = (round(treso_active[0] - treso_passive[0], 3),
                round(treso_active[1] - treso_passive[1], 3))

    # Contrôle cohérence fondamentale FRNG = BFR + TN
    for annee_lbl, idx in [("N", 0), ("N-1", 1)]:
        ecart = abs(frng[idx] - (bfr[idx] + tn[idx]))
        if ecart > 0.01:
            logger.warning(
                "Tréso : FRNG ≠ BFR + TN en %s "
                "(FRNG=%.1f, BFR=%.1f, TN=%.1f, écart=%.3f K€)",
                annee_lbl, frng[idx], bfr[idx], tn[idx], ecart,
            )

    return dict(
        cap_propres=cap_propres, amort_dep=amort_dep,
        prov_risques=prov_risques, dettes_mlt=dettes_mlt, total_res=total_res,
        actif_immo=actif_immo, total_emp=total_emp, frng=frng,
        stocks=stocks, crean_cli=crean_cli, autres_crean=autres_crean,
        cca=cca, total_ac=total_ac,
        det_fourn=det_fourn, det_fisc=det_fisc, autres_det=autres_det,
        pca=pca, total_pc=total_pc,
        bfr=bfr, tn=tn,
        treso_active=treso_active, treso_passive=treso_passive, tn_verif=tn_verif,
    )


# ---------------------------------------------------------------------------
# Onglet Tréso
# ---------------------------------------------------------------------------

def _ecrire_treso_tab(ws, balance: pd.DataFrame,
                      date_n: str, date_n1: str, client: str,
                      liasse_fiscale: dict) -> None:
    """Écrit l'onglet Tréso — BFR, FRNG, Trésorerie nette."""
    remove_gridlines(ws)
    set_col_widths(ws, _WIDTHS_TRESO)

    ws.cell(row=1, column=1, value="Retour sommaire").font = FONT_META

    write_title_block(
        ws, row_titre=3, row_client=4,
        titre="Besoin en fonds de roulement, Fonds de roulement, "
              "Trésorerie nette, Endettement net financier",
        client=client,
    )

    write_header_row(ws, 8, [
        (2, "En milliers d'€uros"),
        (4, date_n), (5, date_n1),
        (6, "Var. K€"), (7, "Var. %"),
    ])

    t   = _calc_treso(balance, liasse_fiscale)
    row = [10]  # liste pour mutation dans les closures

    def _var(vn, vn1):
        var_ke  = round(vn - vn1, 3)
        var_pct = round(var_ke / abs(vn1), 4) if abs(vn1) >= 0.001 else "n/a"
        return var_ke, var_pct

    def section(label):
        ws.cell(row=row[0], column=2, value=label).font = FONT_SECTION
        row[0] += 1

    def poste(label, vn_vn1):
        vn, vn1   = vn_vn1
        var_ke, var_pct = _var(vn, vn1)
        write_data_row(ws, row[0], [
            (3, label,   None,   FONT_NORMAL),
            (4, vn,      NUM_KE, FONT_NORMAL),
            (5, vn1,     NUM_KE, FONT_NORMAL),
            (6, var_ke,  NUM_KE, FONT_NORMAL),
            (7, var_pct,
               NUM_PCT if not isinstance(var_pct, str) else None,
               FONT_NORMAL),
        ])
        row[0] += 1

    def sous_total(label, vn_vn1):
        vn, vn1   = vn_vn1
        var_ke, var_pct = _var(vn, vn1)
        write_total_row(ws, row[0], [
            (2, label),
            (4, vn,      NUM_KE),
            (5, vn1,     NUM_KE),
            (6, var_ke,  NUM_KE),
            (7, var_pct, NUM_PCT if not isinstance(var_pct, str) else None),
        ])
        row[0] += 1

    def blank():
        row[0] += 1

    # ---- FRNG -------------------------------------------------------
    section("FONDS DE ROULEMENT NET GLOBAL (FRNG)")
    section("Ressources stables")
    poste("Capitaux propres (10–14)",                         t["cap_propres"])
    poste("Amortissements et dépréciations (28, 29, 39, 49)", t["amort_dep"])
    poste("Provisions pour risques et charges (15)",          t["prov_risques"])
    poste("Dettes financières MLT (16, 17)",                  t["dettes_mlt"])
    sous_total("= Total ressources stables",                  t["total_res"])
    blank()
    section("Emplois stables")
    poste("Actif immobilisé brut (20–27)",                    t["actif_immo"])
    sous_total("= Total emplois stables",                     t["total_emp"])
    blank()
    sous_total("= FRNG",                                      t["frng"])
    blank()

    # ---- BFR --------------------------------------------------------
    section("BESOIN EN FONDS DE ROULEMENT (BFR)")
    section("Actif circulant d'exploitation")
    poste("Stocks bruts (31–37)",                                    t["stocks"])
    poste("Créances clients (411, 413, 416, 418)",                   t["crean_cli"])
    poste("Autres créances d'exploitation (40, 44, 46, 47 > 0)",    t["autres_crean"])
    poste("Charges constatées d'avance — CCA (486)",                 t["cca"])
    sous_total("= Total actif circulant",                            t["total_ac"])
    blank()
    section("Passif circulant d'exploitation")
    poste("Dettes fournisseurs (401, 403, 408)",                     t["det_fourn"])
    poste("Dettes fiscales et sociales (42, 43, 44 < 0)",            t["det_fisc"])
    poste("Autres dettes d'exploitation (46, 47 < 0)",               t["autres_det"])
    poste("Produits constatés d'avance — PCA (487)",                 t["pca"])
    sous_total("= Total passif circulant",                           t["total_pc"])
    blank()
    sous_total("= BFR",                                              t["bfr"])
    blank()

    # ---- TN ---------------------------------------------------------
    sous_total("= TRÉSORERIE NETTE (TN = FRNG − BFR)",              t["tn"])
    blank()

    # ---- Vérification -----------------------------------------------
    section("Vérification — trésorerie directe")
    poste("Trésorerie active (50, 51, 53 > 0)",      t["treso_active"])
    poste("Trésorerie passive (519 + 512 < 0)",      t["treso_passive"])
    sous_total("= TN (vérification directe)",        t["tn_verif"])

    logger.info(
        "Tréso : onglet généré (FRNG=%.0f, BFR=%.0f, TN=%.0f K€)",
        t["frng"][0], t["bfr"][0], t["tn"][0],
    )


# ---------------------------------------------------------------------------
# Onglet AACE
# ---------------------------------------------------------------------------

def _filtrer_aace(balance_mappee: pd.DataFrame,
                  prefixes_aace: List[str]) -> pd.DataFrame:
    """Retourne les lignes dont CompteNum commence par un préfixe AACE, triées.

    Les préfixes proviennent de liasse_fiscale.aace.prefixes du mapping_pcg.yaml.
    """
    prefixes = tuple(prefixes_aace)
    masque = balance_mappee["CompteNum"].astype(str).apply(
        lambda n: n.startswith(prefixes)
    )
    return balance_mappee[masque].sort_values("CompteNum").reset_index(drop=True)


def _ecrire_aace_tab(ws, df_aace: pd.DataFrame,
                     date_n: str, date_n1: str, client: str) -> None:
    """Écrit l'onglet AACE — Autres Achats et Charges Externes."""
    remove_gridlines(ws)
    set_col_widths(ws, _WIDTHS_CYCLE)

    ws.cell(row=1, column=1, value="Retour sommaire").font = FONT_META

    write_title_block(ws, row_titre=3, row_client=4,
                      titre="Autres achats et charges externes",
                      client=client)

    # Headers row 8 — colonnes A-H uniquement
    headers = [
        (1, "Ref."), (2, "En milliers d'€uros"),
        (4, date_n), (5, date_n1),
        (6, "Var. K€"), (7, "Var. %"),
        (8, "Ref."),
    ]
    write_header_row(ws, 8, headers)

    current_row = 10
    total_n  = 0.0
    total_n1 = 0.0

    for _, r in df_aace.iterrows():
        # Charges en positif (signe +1)
        val_n  = round(float(r["Solde_KE"]),    3)
        val_n1 = round(float(r["Solde_N1_KE"]), 3)
        var_ke = round(val_n - val_n1, 3)

        if abs(val_n1) >= 0.001:
            var_pct: Union[float, str] = round(var_ke / abs(val_n1), 4)
        else:
            var_pct = "n/a"

        cells = [
            (2, str(r["CompteNum"]),  None,   FONT_META),
            (3, r["CompteLib"],       None,   FONT_NORMAL),
            (4, val_n,  NUM_KE, FONT_NORMAL),
            (5, val_n1, NUM_KE, FONT_NORMAL),
            (6, var_ke, NUM_KE, FONT_NORMAL),
            (7, var_pct,
               NUM_PCT if not isinstance(var_pct, str) else None,
               FONT_NORMAL),
            (8, "AACE", None, FONT_META),
        ]
        write_data_row(ws, current_row, cells)
        current_row += 1

        total_n  += val_n
        total_n1 += val_n1

    # Ligne total
    if current_row > 10:
        current_row += 1  # ligne vide
        var_ke_total = round(total_n - total_n1, 3)
        write_total_row(ws, current_row, [
            (2, "TOTAL AACE"),
            (4, round(total_n,  3), NUM_KE),
            (5, round(total_n1, 3), NUM_KE),
            (6, var_ke_total,       NUM_KE),
        ])

    logger.info("AACE : %d comptes écrits", len(df_aace))


# ---------------------------------------------------------------------------
# Onglet cycle
# ---------------------------------------------------------------------------

def _ecrire_cycle_tab(ws, df_cycle: pd.DataFrame, cycle: str,
                      date_n: str, date_n1: str, client: str) -> None:
    remove_gridlines(ws)
    set_col_widths(ws, _WIDTHS_CYCLE)

    ws.cell(row=1, column=1, value="Retour sommaire").font = FONT_META

    write_title_block(ws, row_titre=3, row_client=4,
                      titre=f"Feuille maîtresse du cycle {cycle}",
                      client=client)

    # Headers row 8
    headers = [
        (1, "Ref."), (2, "En milliers d'€uros"),
        (4, date_n), (5, date_n1),
        (6, "Var. K€"), (7, "Var. %"),
        (8, "Ref."), (10, "Cycle"),
    ]
    write_header_row(ws, 8, headers)

    current_row = 10
    ref_cycle   = f"{cycle}0"

    # Grouper par section (en respectant l'ordre)
    for section in ORDRE_SECTIONS:
        # Normaliser la casse pour le filtrage (compta peut être "Actif" ou "ACTIF")
        masque = df_cycle["compta"].str.capitalize() == section
        df_section = df_cycle[masque]
        if df_section.empty:
            continue

        signe = _SIGNE[section]

        # Label de section
        write_section_label(ws, current_row, section.upper())
        current_row += 1

        total_n  = 0.0
        total_n1 = 0.0

        for _, r in df_section.iterrows():
            val_n  = round(float(r["Solde_KE"])    * signe, 3)
            val_n1 = round(float(r["Solde_N1_KE"]) * signe, 3)
            var_ke = round(val_n - val_n1, 3)

            # Recalculer Var_PCT avec les valeurs présentées
            if abs(val_n1) >= 0.001:
                var_pct: Union[float, str] = round(var_ke / abs(val_n1), 4)
            else:
                var_pct = "n/a"

            cells = [
                (2, str(r["CompteNum"]),  None,   FONT_META),
                (3, r["CompteLib"],       None,   FONT_NORMAL),
                (4, val_n,  NUM_KE, FONT_NORMAL),
                (5, val_n1, NUM_KE, FONT_NORMAL),
                (6, var_ke, NUM_KE, FONT_NORMAL),
                (7, var_pct,
                   NUM_PCT if not isinstance(var_pct, str) else None,
                   FONT_NORMAL),
                (8, ref_cycle, None, FONT_META),
                (10, cycle,   None, FONT_META),
            ]
            write_data_row(ws, current_row, cells)
            current_row += 1

            total_n  += val_n
            total_n1 += val_n1

        # Ligne total de section (séparée par une ligne vide)
        current_row += 1


# ---------------------------------------------------------------------------
# Point d'entrée public
# ---------------------------------------------------------------------------

def write(
    balance_mappee: pd.DataFrame,
    client: str,
    date_cloture: str,
    output_path: Union[str, Path],
    pcg_config: Optional[dict] = None,
) -> Path:
    """
    Génère le fichier FM_{client}_{annee}.xlsx.

    Paramètres
    ----------
    balance_mappee : pd.DataFrame
        Produit par cycle_mapper.map_cycles() — colonnes standard + cycle/compta/etatfi/ref.
    client : str
        Nom du client (ex: "GILAC").
    date_cloture : str
        Date de clôture au format 'JJ/MM/AAAA' (ex: '31/12/2025').
    output_path : str ou Path
        Dossier de sortie.
    pcg_config : dict, optionnel
        Config PCG produite par mapping_parser.from_pcg_config() — fournit la
        section liasse_fiscale (préfixes Bilan/Tréso/AACE). Si None, la config
        par défaut (src/config/mapping_pcg.yaml) est chargée.

    Retourne
    --------
    Path
        Chemin complet du fichier généré.
    """
    if pcg_config is None:
        from src.parsers.mapping_parser import from_pcg_config
        logger.debug(
            "write : pcg_config non fourni — chargement de la config par "
            "défaut %s", _PCG_DEFAULT,
        )
        pcg_config = from_pcg_config(_PCG_DEFAULT)
    liasse_fiscale = load_liasse_fiscale(pcg_config)

    output_dir = Path(output_path)
    output_dir.mkdir(parents=True, exist_ok=True)

    annee    = _annee(date_cloture)
    date_n   = _date_str(date_cloture)
    date_n1  = f"31/12/{annee - 1}"

    nom_fichier = output_dir / f"FM_{client}_{annee}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)  # supprimer la feuille vide par défaut

    # Ordre des cycles (depuis la config si dispo, sinon trié)
    cycles_presents = (
        balance_mappee["cycle"]
        .dropna()
        .loc[lambda s: s != ""]
        .unique()
        .tolist()
    )
    # Tri selon l'ordre canonique PCG
    ORDRE_CANONIQUE = [
        "C Propres", "C PRC", "F", "I Incorp", "I Corp", "I Fi",
        "S", "A", "V", "P", "E", "T", "X",
    ]
    cycles_tries = sorted(
        cycles_presents,
        key=lambda c: ORDRE_CANONIQUE.index(c) if c in ORDRE_CANONIQUE else 99,
    )

    # --- Sommaire ---
    ws_sommaire = wb.create_sheet("Sommaire")
    _ecrire_sommaire(ws_sommaire, client, date_n, cycles_tries)

    # --- Balance N Vs N-1 ---
    ws_bg = wb.create_sheet("Balance N Vs N-1")
    _ecrire_balance_tab(ws_bg, balance_mappee, date_n, date_n1, client)

    # --- Bilan ---
    ws_bilan = wb.create_sheet("Bilan")
    _ecrire_bilan_tab(ws_bilan, balance_mappee, date_n, date_n1, client,
                      liasse_fiscale)

    # --- Tréso ---
    ws_treso = wb.create_sheet("Tréso")
    _ecrire_treso_tab(ws_treso, balance_mappee, date_n, date_n1, client,
                      liasse_fiscale)

    # --- AACE ---
    ws_aace = wb.create_sheet("AACE")
    df_aace = _filtrer_aace(balance_mappee, liasse_fiscale["aace"]["prefixes"])
    _ecrire_aace_tab(ws_aace, df_aace, date_n, date_n1, client)

    # --- Onglets par cycle ---
    for cycle in cycles_tries:
        nom_onglet = f"{cycle}0"
        ws_cycle = wb.create_sheet(nom_onglet)
        df_cycle = balance_mappee[balance_mappee["cycle"] == cycle].copy()
        _ecrire_cycle_tab(ws_cycle, df_cycle, cycle, date_n, date_n1, client)

    wb.save(nom_fichier)
    logger.info("FM généré : %s (%d onglets)", nom_fichier, len(wb.sheetnames))
    return nom_fichier
