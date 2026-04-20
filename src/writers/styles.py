"""
Styles Excel centralisés — Design A (minimaliste).

Zéro couleur de fond. Hiérarchie par typographie seule.
"""

from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.styles.numbers import FORMAT_NUMBER
from openpyxl.worksheet.worksheet import Worksheet

# =============================================================================
# Polices
# =============================================================================

FONT_TITLE    = Font(name="Arial", size=16, bold=True)
FONT_SUBTITLE = Font(name="Arial", size=11, italic=True, color="808080")
FONT_HEADER   = Font(name="Arial", size=10, bold=True,   color="808080")
FONT_NORMAL   = Font(name="Arial", size=10)
FONT_BOLD     = Font(name="Arial", size=10, bold=True)
FONT_SECTION  = Font(name="Arial", size=9,  bold=True,   color="999999")
FONT_META     = Font(name="Arial", size=9,  color="AAAAAA")

# =============================================================================
# Bordures
# =============================================================================

_SIDE_MED   = Side(style="medium", color="000000")
_SIDE_HAIR  = Side(style="hair",   color="C0C0C0")
_SIDE_THIN  = Side(style="thin",   color="000000")

BORDER_BOTTOM_MED  = Border(bottom=_SIDE_MED)
BORDER_BOTTOM_HAIR = Border(bottom=_SIDE_HAIR)
BORDER_TOP_THIN    = Border(top=_SIDE_THIN)

# =============================================================================
# Formats numériques
# =============================================================================

NUM_KE  = '#,##0;(#,##0);"-"'   # entiers, négatifs entre parenthèses, zéros = "-"
NUM_PCT = "0%"                   # pourcentages sans décimales

# =============================================================================
# Fonctions utilitaires
# =============================================================================


def remove_gridlines(ws: Worksheet) -> None:
    """Supprime le quadrillage de la feuille."""
    ws.sheet_view.showGridLines = False


def set_col_widths(ws: Worksheet, widths: dict) -> None:
    """
    Applique les largeurs de colonnes.

    widths : dict {lettre_colonne: largeur}  ex. {"A": 6, "B": 12}
    """
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _apply(cell, font=None, border=None, number_format=None, alignment=None):
    """Applique les attributs fournis à une cellule."""
    if font is not None:
        cell.font = font
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format
    if alignment is not None:
        cell.alignment = alignment


def write_title_block(ws: Worksheet, row_titre: int, row_client: int,
                      titre: str, client: str) -> None:
    """Écrit le bloc titre + client (lignes 3-4 dans les onglets FM)."""
    c_titre  = ws.cell(row=row_titre,  column=4, value=titre)
    c_client = ws.cell(row=row_client, column=4, value=client)
    _apply(c_titre,  font=FONT_TITLE)
    _apply(c_client, font=FONT_SUBTITLE)


def write_header_row(ws: Worksheet, row: int, headers: list) -> None:
    """
    Écrit une ligne d'en-têtes avec trait épais en dessous.

    headers : liste de (colonne_index_1based, valeur)
    """
    for col, value in headers:
        c = ws.cell(row=row, column=col, value=value)
        _apply(c, font=FONT_HEADER, border=BORDER_BOTTOM_MED)


def write_data_row(ws: Worksheet, row: int, cells: list) -> None:
    """
    Écrit une ligne de données avec trait fin en dessous.

    cells : liste de (colonne_index_1based, valeur, format_optionnel)
    """
    for item in cells:
        col, value = item[0], item[1]
        fmt = item[2] if len(item) > 2 else None
        font = item[3] if len(item) > 3 else FONT_NORMAL
        c = ws.cell(row=row, column=col, value=value)
        _apply(c, font=font, border=BORDER_BOTTOM_HAIR, number_format=fmt)


def write_section_label(ws: Worksheet, row: int, label: str) -> None:
    """Écrit un label de section ACTIF/PASSIF/CHARGES/PRODUITS."""
    c = ws.cell(row=row, column=2, value=label)
    _apply(c, font=FONT_SECTION)


def write_total_row(ws: Worksheet, row: int, cells: list) -> None:
    """
    Écrit une ligne total avec trait au-dessus et police bold.

    cells : liste de (colonne_index_1based, valeur, format_optionnel)
    """
    for item in cells:
        col, value = item[0], item[1]
        fmt = item[2] if len(item) > 2 else None
        c = ws.cell(row=row, column=col, value=value)
        _apply(c, font=FONT_BOLD, border=BORDER_TOP_THIN, number_format=fmt)
