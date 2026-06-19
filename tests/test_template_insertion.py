"""
Tests de l'insertion des feuilles FM comme onglets dans les templates (P1)
et de la préservation des images/dessins (9 des 13 templates en contiennent).

Utilise un FM synthétique (mêmes noms d'onglets qu'un FM réel, valeurs et
styles connus) : l'insertion ne dépend que des noms d'onglets, et la
fidélité de copie (valeurs, styles, fusions) est vérifiable précisément.
"""

import io
import zipfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl import Workbook
from openpyxl.styles import Font

from src.parsers.mapping_parser import from_pcg_config
from src.writers.template_writer import write
from src.writers.worksheet_copy import copy_worksheet

DATA_DIR      = Path(__file__).parent.parent / "data"
TEMPLATES_DIR = DATA_DIR / "templates"
CONFIG_DIR    = Path(__file__).parent.parent / "src" / "config"

CLIENT       = "GILAC"
DATE_CLOTURE = "31/12/2025"

FM_SHEETNAMES = [
    "Sommaire", "Balance N Vs N-1", "Bilan", "EBIT",
    "Actif détaillé", "Passif détaillé", "P&L détaillé", "Tréso", "AACE",
    "C Propres0", "C PRC0", "F0", "I Incorp0", "I Corp0", "I Fi0",
    "S0", "A0", "V0", "P0", "E0", "T0", "X0",
]

pytestmark = pytest.mark.skipif(
    not TEMPLATES_DIR.exists(),
    reason="Templates absents (data/ non committé)",
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def pcg():
    return from_pcg_config(CONFIG_DIR / "mapping_pcg.yaml")


@pytest.fixture(scope="module")
def fm_path(tmp_path_factory) -> Path:
    """FM synthétique : valeurs, style et fusion connus dans chaque onglet."""
    chemin = tmp_path_factory.mktemp("fm_synthetique") / "FM_GILAC_2025.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    for nom in FM_SHEETNAMES:
        ws = wb.create_sheet(nom)
        ws["B2"] = f"Feuille {nom}"
        ws["B2"].font = Font(name="Arial", size=10, bold=True)
        ws["D10"] = 123.456
        ws["D10"].number_format = '#,##0;(#,##0);"-"'
        ws.merge_cells("B2:C2")
        ws.column_dimensions["B"].width = 42
        ws.sheet_view.showGridLines = False
    wb.save(chemin)
    return chemin


@pytest.fixture(scope="module")
def zip_contents(pcg, fm_path, tmp_path_factory) -> dict:
    """Génère le ZIP avec insertion FM et retourne {nom: bytes}."""
    out = tmp_path_factory.mktemp("templates_insertion")
    zip_path = write(TEMPLATES_DIR, CLIENT, DATE_CLOTURE, out,
                     pcg["templates"], fm_path=fm_path,
                     integration_templates=pcg["integration_templates"])
    with zipfile.ZipFile(zip_path) as zf:
        return {name: zf.read(name) for name in zf.namelist()}


def _wb(zip_contents: dict, nom: str):
    return openpyxl.load_workbook(io.BytesIO(zip_contents[nom]))


# ---------------------------------------------------------------------------
# Tests : position et contenu des onglets insérés
# ---------------------------------------------------------------------------

def test_a_contient_aace_et_a0_avant_synthese(zip_contents):
    """Le template A contient AACE puis A0, juste avant 'A Synthèse'."""
    wb = _wb(zip_contents, "2025_12_GILAC_A_Fournisseurs-Achats.xlsx")
    noms = wb.sheetnames
    assert "AACE" in noms and "A0" in noms
    idx_synth = noms.index("A Synthèse")
    assert noms.index("AACE") == idx_synth - 2
    assert noms.index("A0") == idx_synth - 1


def test_f_contient_treso_f0_et_i_corp_les_deux_immos(zip_contents):
    """F reçoit [Tréso, F0] ; I corp reçoit I Corp0 ET I Incorp0 (confirmé
    cabinet : les deux immos dans le même fichier template)."""
    wb_f = _wb(zip_contents, "2025_12_GILAC_F_Financement.xlsx")
    noms_f = wb_f.sheetnames
    idx_synth = next(i for i, n in enumerate(noms_f) if "synth" in n.lower())
    assert noms_f.index("Tréso") == idx_synth - 2
    assert noms_f.index("F0") == idx_synth - 1

    wb_i = _wb(zip_contents, "2025_12_GILAC_I_Immobilisations corp.xlsx")
    noms_i = wb_i.sheetnames
    idx_synth_i = next(i for i, n in enumerate(noms_i) if "synth" in n.lower())
    assert noms_i.index("I Corp0") == idx_synth_i - 2
    assert noms_i.index("I Incorp0") == idx_synth_i - 1


def test_i_incorporelles_aucune_feuille_inseree(zip_contents):
    """Le template I incorporelles ne reçoit AUCUNE feuille FM (copie +
    renommage + placeholders uniquement) — arbitrage figé 09/06/2026."""
    wb = _wb(zip_contents, "2025_12_GILAC_I_Immobilisations incorporelles.xlsx")
    assert not any(nom in wb.sheetnames for nom in FM_SHEETNAMES)
    # Structure d'origine inchangée (2 feuilles)
    assert len(wb.sheetnames) == 2


def test_valeurs_styles_et_fusions_copies(zip_contents):
    """L'onglet inséré reproduit valeurs, police, format et fusions du FM."""
    wb = _wb(zip_contents, "2025_12_GILAC_V_Clients-Ventes.xlsx")
    ws = wb["V0"]
    assert ws["B2"].value == "Feuille V0"
    assert ws["B2"].font.bold is True
    assert ws["B2"].font.name == "Arial"
    assert ws["D10"].value == 123.456
    assert ws["D10"].number_format == '#,##0;(#,##0);"-"'
    assert "B2:C2" in [str(r) for r in ws.merged_cells.ranges]
    assert ws.column_dimensions["B"].width == 42


# ---------------------------------------------------------------------------
# Test : préservation des images / dessins (obligatoire — P1)
# ---------------------------------------------------------------------------

def _compter_medias(data: bytes):
    """Compte les entrées xl/media/* et xl/drawings/* d'une archive xlsx."""
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        noms = zf.namelist()
    media = [n for n in noms if n.startswith("xl/media/")]
    drawings = [n for n in noms if n.startswith("xl/drawings/")
                and not n.startswith("xl/drawings/_rels/")]
    return len(media), len(drawings)


def test_aucune_perte_media_ni_dessin(zip_contents):
    """Chaque fichier généré contient AU MOINS autant d'entrées xl/media/*
    et xl/drawings/* que son template source (openpyxl en perd à la simple
    ouverture+sauvegarde : la restauration ZIP/XML doit tout réinjecter)."""
    prefixe_sortie = f"2025_12_{CLIENT}_"
    for nom, data in zip_contents.items():
        source = TEMPLATES_DIR / nom.replace(prefixe_sortie, "20XX_XX_YYY_")
        assert source.exists(), f"Template source introuvable pour {nom}"
        media_src, dessins_src = _compter_medias(source.read_bytes())
        media_out, dessins_out = _compter_medias(data)
        assert media_out >= media_src, (
            f"{nom} : {media_src - media_out} média(s) perdu(s) "
            f"({media_src} → {media_out})"
        )
        assert dessins_out >= dessins_src, (
            f"{nom} : {dessins_src - dessins_out} dessin(s) perdu(s) "
            f"({dessins_src} → {dessins_out})"
        )


def test_fichiers_generes_relisibles(zip_contents):
    """Tous les fichiers générés se rechargent sans erreur (XML valide
    après la restauration des dessins)."""
    for nom, data in zip_contents.items():
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert wb.sheetnames, nom


# ---------------------------------------------------------------------------
# Tests unitaires : copy_worksheet
# ---------------------------------------------------------------------------

def test_copy_worksheet_position(fm_path):
    """copy_worksheet insère à l'index demandé."""
    wb_src = openpyxl.load_workbook(fm_path)
    wb_dst = Workbook()
    wb_dst.active.title = "Synthèse"
    copy_worksheet(wb_dst, wb_src["A0"], index=0)
    assert wb_dst.sheetnames == ["A0", "Synthèse"]


def test_copy_worksheet_refuse_doublon(fm_path):
    """copy_worksheet lève ValueError si le nom d'onglet existe déjà."""
    wb_src = openpyxl.load_workbook(fm_path)
    wb_dst = Workbook()
    wb_dst.active.title = "A0"
    with pytest.raises(ValueError, match="existe déjà"):
        copy_worksheet(wb_dst, wb_src["A0"])
