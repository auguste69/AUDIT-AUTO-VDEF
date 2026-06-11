"""
Tests d'intégration de l'onglet P&L détaillé sur le jeu de données GILAC.

Valeur cible GILAC 2025 (validée par le cabinet) :
- Résultat net ≈ 2 572,7 K€ (tolérance 5 K€) — égal à −(somme classes 6/7)

Le FEC n'est parsé qu'une fois (fixtures module, même pattern que
tests/test_ebit_gilac.py).
"""

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from src.engine.balance_builder import build
from src.engine.cycle_mapper import map_cycles
from src.engine.financial_engine import calculer_pl_detaille
from src.engine.liasse_fiscale_loader import load_liasse_fiscale
from src.parsers.fec_parser import parse
from src.parsers.mapping_parser import from_fm, from_pcg_config
from src.writers.fm_writer import write

DATA_DIR   = Path(__file__).parent.parent / "data"
CONFIG_DIR = Path(__file__).parent.parent / "src" / "config"

FEC_PATH = DATA_DIR / "GILAC_2025_12_31_FEC.txt"
FM_REF   = DATA_DIR / "FM GILAC.xlsx"
PCG_PATH = CONFIG_DIR / "mapping_pcg.yaml"

RESULTAT_NET_CIBLE_KE = 2572.7

pytestmark = pytest.mark.skipif(
    not FEC_PATH.exists() or not FM_REF.exists(),
    reason="Fichiers GILAC absents (data/ non committé)",
)


# ---------------------------------------------------------------------------
# Fixtures (module : le FEC n'est parsé qu'une seule fois)
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def balance_mappee() -> pd.DataFrame:
    """Construit la balance mappée GILAC avec soldes N-1 (pipeline complet)."""
    df  = parse(FEC_PATH)
    mfm = from_fm(FM_REF)
    pcg = from_pcg_config(PCG_PATH)

    wb = openpyxl.load_workbook(FM_REF, read_only=True, data_only=True)
    ws = wb["Balance N Vs N-1"]
    balance_n1: dict = {}
    for row in ws.iter_rows(min_row=10, values_only=True):
        if row[1] is None:
            continue
        try:
            num = str(int(float(row[1])))
        except (ValueError, TypeError):
            continue
        solde = float(row[3]) if row[3] is not None else 0.0
        balance_n1[num] = {
            "libelle":  str(row[2]) if row[2] else "",
            "solde_ke": solde,
        }
    wb.close()

    bal = build(df, balance_n1)
    return map_cycles(bal, mfm, pcg)


@pytest.fixture(scope="module")
def liasse(balance_mappee) -> dict:
    """Config liasse_fiscale complète (ebit + pl_detaille + structures)."""
    return load_liasse_fiscale(from_pcg_config(PCG_PATH))


@pytest.fixture(scope="module")
def pl_gilac(balance_mappee, liasse):
    """Calcule le P&L détaillé GILAC."""
    return calculer_pl_detaille(balance_mappee, liasse)


@pytest.fixture(scope="module")
def fm_sheetnames(balance_mappee, tmp_path_factory) -> list:
    """Génère le FM complet et retourne la liste ordonnée de ses onglets."""
    out = tmp_path_factory.mktemp("pl_detaille_gilac")
    pcg = from_pcg_config(PCG_PATH)
    fm_path = write(balance_mappee, "GILAC", "31/12/2025", out, pcg_config=pcg)
    wb = openpyxl.load_workbook(fm_path, read_only=True)
    noms = list(wb.sheetnames)
    wb.close()
    return noms


# ---------------------------------------------------------------------------
# Tests : valeurs cibles
# ---------------------------------------------------------------------------

def test_resultat_net_cible(pl_gilac):
    """Résultat net 2025 ≈ 2 572,7 K€ (tolérance 5 K€)."""
    assert abs(pl_gilac.resultat_net.valeur_n - RESULTAT_NET_CIBLE_KE) <= 5.0, (
        f"Résultat net N = {pl_gilac.resultat_net.valeur_n:.1f} K€, "
        f"cible {RESULTAT_NET_CIBLE_KE} K€"
    )


def test_resultat_net_egal_moins_somme_classes_67(pl_gilac, balance_mappee):
    """Couverture totale : résultat net = −(somme des classes 6 et 7)."""
    masque = balance_mappee["CompteNum"].astype(str).str.startswith(("6", "7"))
    attendu = -float(balance_mappee.loc[masque, "Solde_KE"].sum())
    assert abs(pl_gilac.resultat_net.valeur_n - attendu) <= 0.01


def test_coherence_resultats_intermediaires(pl_gilac):
    """Résultat net = courant + exceptionnel + participation + impôts."""
    d = pl_gilac.as_dict()
    recompose = (d["resultat_courant"][0] + d["resultat_exceptionnel"][0]
                 + d["participation_salaries"][0] + d["impots_benefices"][0])
    assert abs(d["resultat_net"][0] - recompose) <= 0.01


def test_resultat_financier(pl_gilac):
    """Résultat financier = produits financiers + charges financières."""
    d = pl_gilac.as_dict()
    assert abs(
        d["resultat_financier"][0]
        - (d["produits_financiers"][0] + d["charges_financieres"][0])
    ) <= 0.01
    # Les charges financières sont présentées en négatif (convention liasse)
    assert d["charges_financieres"][0] <= 0.0


def test_exploitation_residuelle_vs_ebit(pl_gilac):
    """Exploitation = postes EBIT + résiduels divers (comptes hors EBIT)."""
    d = pl_gilac.as_dict()
    assert abs(
        d["produits_exploitation"][0]
        - (pl_gilac.ebit.total_produits.valeur_n + d["produits_expl_divers"][0])
    ) <= 0.01
    assert abs(
        d["charges_exploitation"][0]
        - (-pl_gilac.ebit.total_charges.valeur_n + d["charges_expl_divers"][0])
    ) <= 0.01


# ---------------------------------------------------------------------------
# Tests : onglet P&L détaillé dans le FM généré
# ---------------------------------------------------------------------------

def test_onglet_pl_detaille_existe(fm_sheetnames):
    """L'onglet P&L détaillé existe dans le FM généré."""
    assert "P&L détaillé" in fm_sheetnames


def test_ordre_onglets_cerfa(fm_sheetnames):
    """Ordre : Bilan → EBIT → Actif détaillé → Passif détaillé →
    P&L détaillé → Tréso → AACE."""
    attendu = ["Bilan", "EBIT", "Actif détaillé", "Passif détaillé",
               "P&L détaillé", "Tréso", "AACE"]
    indices = [fm_sheetnames.index(nom) for nom in attendu]
    assert indices == sorted(indices)
    assert indices[-1] - indices[0] == len(attendu) - 1  # consécutifs
