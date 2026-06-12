"""
Test de régression cellulaire du FM GILAC.

Régénère le FM complet dans un répertoire temporaire et le compare, cellule
par cellule (valeurs + number_format), à la fixture de référence
tests/fixtures/FM_GILAC_2025_POST_FIX.xlsx (snapshot pris après les
corrections d'audit : comptes N-1 soldés inclus dans la balance, Tréso
cohérente — résultat en cours et classe 4/5 par signe, préfixes Bilan
512/517/519/509 sans double comptage).

Règles de comparaison :
- Comparaison PAR NOM D'ONGLET (jamais par index) — uniquement les onglets
  présents dans la fixture ; les onglets supplémentaires du FM régénéré
  sont ignorés.
- Flottants : tolérance abs(a - b) <= 0.01.
- Pas de comparaison octet par octet.
"""

from pathlib import Path
from typing import Iterator, Tuple

import openpyxl
import pandas as pd
import pytest

from src.parsers.fec_parser import parse
from src.parsers.mapping_parser import from_fm, from_pcg_config
from src.engine.balance_builder import build
from src.engine.cycle_mapper import map_cycles
from src.writers.fm_writer import write

DATA_DIR    = Path(__file__).parent.parent / "data"
CONFIG_DIR  = Path(__file__).parent.parent / "src" / "config"
FIXTURE_FM  = Path(__file__).parent / "fixtures" / "FM_GILAC_2025_POST_FIX.xlsx"

FEC_PATH = DATA_DIR / "GILAC_2025_12_31_FEC.txt"
FM_REF   = DATA_DIR / "FM GILAC.xlsx"
PCG_PATH = CONFIG_DIR / "mapping_pcg.yaml"

_TOLERANCE = 0.01

pytestmark = pytest.mark.skipif(
    not FEC_PATH.exists() or not FM_REF.exists(),
    reason="Fichiers GILAC absents (data/ non committé)",
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def balance_mappee() -> pd.DataFrame:
    """Construit la balance mappée GILAC (même pipeline que main.py)."""
    df  = parse(FEC_PATH)
    mfm = from_fm(FM_REF)
    pcg = from_pcg_config(PCG_PATH)

    # Soldes N-1 extraits du FM existant (même logique que run_pipeline)
    wb = openpyxl.load_workbook(FM_REF, read_only=True, data_only=True)
    ws = wb["Balance N Vs N-1"]
    balance_n1: dict = {}
    for row in ws.iter_rows(min_row=10, values_only=True):
        if row[1] is None:
            continue
        try:
            num = str(int(float(row[1])))
        except (ValueError, TypeError):
            continue
        solde = float(row[3]) if row[3] is not None else 0.0
        balance_n1[num] = {
            "libelle":  str(row[2]) if row[2] else "",
            "solde_ke": solde,
        }
    wb.close()

    bal = build(df, balance_n1)
    return map_cycles(bal, mfm, pcg)


@pytest.fixture(scope="module")
def fm_regen_path(balance_mappee, tmp_path_factory) -> Path:
    """Régénère le FM dans un tmp_path pytest avec la config PCG par défaut."""
    out = tmp_path_factory.mktemp("fm_regression")
    pcg = from_pcg_config(PCG_PATH)
    return write(balance_mappee, "GILAC", "31/12/2025", out, pcg_config=pcg)


@pytest.fixture(scope="module")
def wb_regen(fm_regen_path):
    return openpyxl.load_workbook(fm_regen_path)


@pytest.fixture(scope="module")
def wb_fixture():
    return openpyxl.load_workbook(FIXTURE_FM)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _valeurs_egales(a, b) -> bool:
    """Égalité de valeurs de cellules — tolérance 0.01 sur les flottants."""
    if isinstance(a, (int, float)) and isinstance(b, (int, float)) \
            and not isinstance(a, bool) and not isinstance(b, bool):
        return abs(float(a) - float(b)) <= _TOLERANCE
    return a == b


def _iter_cellules(ws_fix, ws_new) -> Iterator[Tuple[int, int]]:
    """Itère sur la plage couvrant les deux feuilles (union des dimensions)."""
    max_row = max(ws_fix.max_row, ws_new.max_row)
    max_col = max(ws_fix.max_column, ws_new.max_column)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            yield r, c


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

def test_fixture_presente():
    """La fixture de référence existe."""
    assert FIXTURE_FM.exists(), f"Fixture absente : {FIXTURE_FM}"


def test_onglets_fixture_presents(wb_fixture, wb_regen):
    """Chaque onglet de la fixture existe dans le FM régénéré (par NOM)."""
    manquants = [n for n in wb_fixture.sheetnames if n not in wb_regen.sheetnames]
    assert not manquants, (
        f"Onglets de la fixture absents du FM régénéré : {manquants}. "
        f"Onglets régénérés : {wb_regen.sheetnames}"
    )


def test_regression_cellulaire(wb_fixture, wb_regen):
    """Toutes les cellules (valeur + number_format) sont identiques.

    Comparaison par nom d'onglet, uniquement les onglets de la fixture ;
    les onglets supplémentaires du FM régénéré sont ignorés.
    """
    ecarts = []
    for nom in wb_fixture.sheetnames:
        if nom not in wb_regen.sheetnames:
            continue  # déjà couvert par test_onglets_fixture_presents
        ws_fix = wb_fixture[nom]
        ws_new = wb_regen[nom]
        for r, c in _iter_cellules(ws_fix, ws_new):
            cell_fix = ws_fix.cell(row=r, column=c)
            cell_new = ws_new.cell(row=r, column=c)
            if not _valeurs_egales(cell_fix.value, cell_new.value):
                ecarts.append(
                    f"{nom}!{cell_fix.coordinate} : valeur "
                    f"{cell_fix.value!r} ≠ {cell_new.value!r}"
                )
            elif cell_fix.number_format != cell_new.number_format:
                ecarts.append(
                    f"{nom}!{cell_fix.coordinate} : format "
                    f"{cell_fix.number_format!r} ≠ {cell_new.number_format!r}"
                )
            if len(ecarts) >= 50:  # limiter le volume du rapport d'échec
                break
        if len(ecarts) >= 50:
            break

    assert not ecarts, (
        f"{len(ecarts)} écart(s) cellulaire(s) (50 max affichés) :\n"
        + "\n".join(ecarts)
    )
