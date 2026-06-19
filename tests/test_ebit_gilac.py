"""
Tests d'intégration de l'onglet EBIT sur le jeu de données GILAC.

Valeurs cibles GILAC 2025 (validées par le cabinet) :
- CA ≈ 19 126,6 K€ (tolérance 1 K€)
- EBIT (résultat d'exploitation) ≈ 2 678 K€ (tolérance 5 K€)

Le FEC n'est parsé qu'une fois (fixtures module, même pattern que
tests/test_fm_regression_cellulaire.py).
"""

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from src.engine.balance_builder import build
from src.engine.cycle_mapper import map_cycles
from src.engine.financial_engine import calculer_ebit
from src.engine.liasse_fiscale_loader import load_liasse_fiscale
from src.models.financial_statements import EbitSynthetique
from src.parsers.fec_parser import parse
from src.parsers.mapping_parser import from_fm, from_pcg_config
from src.writers.fm_writer import write

DATA_DIR   = Path(__file__).parent.parent / "data"
CONFIG_DIR = Path(__file__).parent.parent / "src" / "config"

FEC_PATH = DATA_DIR / "GILAC_2025_12_31_FEC.txt"
FM_REF   = DATA_DIR / "FM GILAC.xlsx"
PCG_PATH = CONFIG_DIR / "mapping_pcg.yaml"

CA_CIBLE_KE   = 19126.6
EBIT_CIBLE_KE = 2678.0

pytestmark = pytest.mark.skipif(
    not FEC_PATH.exists() or not FM_REF.exists(),
    reason="Fichiers GILAC absents (data/ non committé)",
)


# ---------------------------------------------------------------------------
# Fixtures (module : le FEC n'est parsé qu'une seule fois)
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def balance_mappee() -> pd.DataFrame:
    """Construit la balance mappée GILAC avec soldes N-1 (pipeline complet)."""
    df  = parse(FEC_PATH)
    mfm = from_fm(FM_REF)
    pcg = from_pcg_config(PCG_PATH)

    # Soldes N-1 extraits du FM existant (même logique que run_pipeline)
    wb = openpyxl.load_workbook(FM_REF, read_only=True, data_only=True)
    ws = wb["Balance N Vs N-1"]
    balance_n1: dict = {}
    for row in ws.iter_rows(min_row=10, values_only=True):
        if row[1] is None:
            continue
        try:
            num = str(int(float(row[1])))
        except (ValueError, TypeError):
            continue
        solde = float(row[3]) if row[3] is not None else 0.0
        balance_n1[num] = {
            "libelle":  str(row[2]) if row[2] else "",
            "solde_ke": solde,
        }
    wb.close()

    bal = build(df, balance_n1)
    return map_cycles(bal, mfm, pcg)


@pytest.fixture(scope="module")
def ebit_gilac(balance_mappee) -> EbitSynthetique:
    """Calcule l'EBIT GILAC depuis la section liasse_fiscale du YAML."""
    pcg = from_pcg_config(PCG_PATH)
    liasse = load_liasse_fiscale(pcg)
    liasse["ebit"] = pcg["liasse_fiscale"]["ebit"]
    return calculer_ebit(balance_mappee, liasse)


@pytest.fixture(scope="module")
def fm_sheetnames(balance_mappee, tmp_path_factory) -> list:
    """Génère le FM complet et retourne la liste ordonnée de ses onglets."""
    out = tmp_path_factory.mktemp("ebit_gilac")
    pcg = from_pcg_config(PCG_PATH)
    fm_path = write(balance_mappee, "GILAC", "31/12/2025", out, pcg_config=pcg)
    wb = openpyxl.load_workbook(fm_path, read_only=True)
    noms = list(wb.sheetnames)
    wb.close()
    return noms


# ---------------------------------------------------------------------------
# Tests : valeurs cibles
# ---------------------------------------------------------------------------

def test_ca_n_cible(ebit_gilac):
    """CA 2025 ≈ 19 126,6 K€ (tolérance 1 K€)."""
    assert abs(ebit_gilac.ca.valeur_n - CA_CIBLE_KE) <= 1.0, (
        f"CA N = {ebit_gilac.ca.valeur_n:.1f} K€, cible {CA_CIBLE_KE} K€"
    )


def test_ebit_n_cible(ebit_gilac):
    """EBIT 2025 ≈ 2 678 K€ (tolérance 5 K€)."""
    assert abs(ebit_gilac.ebit.valeur_n - EBIT_CIBLE_KE) <= 5.0, (
        f"EBIT N = {ebit_gilac.ebit.valeur_n:.1f} K€, cible {EBIT_CIBLE_KE} K€"
    )


def test_coherence_ebit_produits_charges(ebit_gilac):
    """Cohérence interne : EBIT = total produits − total charges (N et N-1)."""
    assert abs(
        ebit_gilac.ebit.valeur_n
        - (ebit_gilac.total_produits.valeur_n
           - ebit_gilac.total_charges.valeur_n)
    ) < 1e-6
    assert abs(
        ebit_gilac.ebit.valeur_n1
        - (ebit_gilac.total_produits.valeur_n1
           - ebit_gilac.total_charges.valeur_n1)
    ) < 1e-6


def test_ca_somme_ventes_production(ebit_gilac):
    """CA = ventes de marchandises + production vendue."""
    d = ebit_gilac.as_dict()
    assert abs(
        d["ca"][0] - (d["ventes_marchandises"][0] + d["production_vendue"][0])
    ) < 1e-6


# ---------------------------------------------------------------------------
# Tests : onglet EBIT dans le FM généré
# ---------------------------------------------------------------------------

def test_onglet_ebit_existe(fm_sheetnames):
    """L'onglet EBIT existe dans le FM généré."""
    assert "EBIT" in fm_sheetnames


def test_onglet_ebit_entre_bilan_et_treso(fm_sheetnames):
    """L'onglet EBIT est positionné entre Bilan et Tréso (les onglets
    cerfa détaillés du prompt 10 s'intercalent entre EBIT et Tréso)."""
    assert fm_sheetnames.index("EBIT") == fm_sheetnames.index("Bilan") + 1
    assert fm_sheetnames.index("EBIT") < fm_sheetnames.index("Tréso")
