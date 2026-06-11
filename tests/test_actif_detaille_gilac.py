"""
Tests d'intégration de l'onglet Actif détaillé (cerfa 2050) — données GILAC.

La partition des structures détaillées reprend exactement les listes du
bilan synthétique : le total de l'Actif détaillé doit être strictement
égal au Total Actif du bilan (tolérance 1 K€ pour les arrondis).

Le FEC n'est parsé qu'une fois (fixtures module).
"""

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from src.engine.balance_builder import build
from src.engine.cycle_mapper import map_cycles
from src.engine.financial_engine import calculer_actif_detaille, calculer_bilan
from src.engine.liasse_fiscale_loader import load_liasse_fiscale
from src.parsers.fec_parser import parse
from src.parsers.mapping_parser import from_fm, from_pcg_config
from src.writers.fm_writer import write

DATA_DIR   = Path(__file__).parent.parent / "data"
CONFIG_DIR = Path(__file__).parent.parent / "src" / "config"

FEC_PATH = DATA_DIR / "GILAC_2025_12_31_FEC.txt"
FM_REF   = DATA_DIR / "FM GILAC.xlsx"
PCG_PATH = CONFIG_DIR / "mapping_pcg.yaml"

pytestmark = pytest.mark.skipif(
    not FEC_PATH.exists() or not FM_REF.exists(),
    reason="Fichiers GILAC absents (data/ non committé)",
)


# ---------------------------------------------------------------------------
# Fixtures (module : le FEC n'est parsé qu'une seule fois)
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def balance_mappee() -> pd.DataFrame:
    """Construit la balance mappée GILAC avec soldes N-1 (pipeline complet)."""
    df  = parse(FEC_PATH)
    mfm = from_fm(FM_REF)
    pcg = from_pcg_config(PCG_PATH)

    wb = openpyxl.load_workbook(FM_REF, read_only=True, data_only=True)
    ws = wb["Balance N Vs N-1"]
    balance_n1: dict = {}
    for row in ws.iter_rows(min_row=10, values_only=True):
        if row[1] is None:
            continue
        try:
            num = str(int(float(row[1])))
        except (ValueError, TypeError):
            continue
        solde = float(row[3]) if row[3] is not None else 0.0
        balance_n1[num] = {
            "libelle":  str(row[2]) if row[2] else "",
            "solde_ke": solde,
        }
    wb.close()

    bal = build(df, balance_n1)
    return map_cycles(bal, mfm, pcg)


@pytest.fixture(scope="module")
def liasse() -> dict:
    """Config liasse_fiscale complète (bilan + structures détaillées)."""
    return load_liasse_fiscale(from_pcg_config(PCG_PATH))


@pytest.fixture(scope="module")
def actif_gilac(balance_mappee, liasse):
    """Calcule l'Actif détaillé GILAC."""
    return calculer_actif_detaille(balance_mappee, liasse)


@pytest.fixture(scope="module")
def bilan_gilac(balance_mappee, liasse):
    """Calcule le bilan synthétique GILAC (référence de cohérence)."""
    return calculer_bilan(balance_mappee, liasse)


@pytest.fixture(scope="module")
def fm_ws(balance_mappee, tmp_path_factory):
    """Génère le FM complet et retourne (sheetnames, onglet Actif détaillé)."""
    out = tmp_path_factory.mktemp("actif_detaille_gilac")
    pcg = from_pcg_config(PCG_PATH)
    fm_path = write(balance_mappee, "GILAC", "31/12/2025", out, pcg_config=pcg)
    wb = openpyxl.load_workbook(fm_path, read_only=True)
    yield wb.sheetnames, wb["Actif détaillé"]
    wb.close()


# ---------------------------------------------------------------------------
# Tests : cohérence avec le bilan synthétique
# ---------------------------------------------------------------------------

def test_total_actif_egal_bilan(actif_gilac, bilan_gilac):
    """Total Actif détaillé = Total Actif du bilan (N et N-1, ± 1 K€)."""
    assert abs(actif_gilac.total.valeur_n
               - bilan_gilac.total_actif.valeur_n) <= 1.0
    assert abs(actif_gilac.total.valeur_n1
               - bilan_gilac.total_actif.valeur_n1) <= 1.0


def test_sous_totaux_immobilisations_egaux_bilan(actif_gilac, bilan_gilac):
    """Les sous-totaux immo incorp/corp/fi raffinent les postes du bilan."""
    sous_totaux = {s.cle: s.sous_total for s in actif_gilac.sections
                   if s.sous_total is not None}
    for cle in ("immo_incorp", "immo_corp", "immo_fi", "stocks"):
        assert abs(sous_totaux[cle].valeur_n
                   - bilan_gilac.postes[cle].valeur_n) <= 0.01, cle


def test_agregat_actif_immobilise(actif_gilac):
    """Actif immobilisé net = immo incorp + immo corp + immo fi."""
    sous_totaux = {s.cle: s.sous_total for s in actif_gilac.sections
                   if s.sous_total is not None}
    agregats = {a.cle: a for a in actif_gilac.agregats}
    attendu = sum(sous_totaux[c].valeur_n
                  for c in ("immo_incorp", "immo_corp", "immo_fi"))
    assert abs(agregats["actif_immobilise"].valeur_n - attendu) <= 0.01


def test_total_egal_somme_des_postes(actif_gilac):
    """Le total est la somme de tous les postes de toutes les sections."""
    somme = sum(p.valeur_n for s in actif_gilac.sections for p in s.postes)
    assert abs(actif_gilac.total.valeur_n - somme) <= 0.01


# ---------------------------------------------------------------------------
# Tests : onglet Actif détaillé dans le FM généré
# ---------------------------------------------------------------------------

def test_onglet_existe(fm_ws):
    """L'onglet Actif détaillé existe dans le FM généré."""
    sheetnames, _ = fm_ws
    assert "Actif détaillé" in sheetnames


def test_headers_dates_en_string(fm_ws):
    """Les dates des en-têtes sont des strings (Design A, pas de datetime)."""
    _, ws = fm_ws
    row8 = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]
    assert row8[3] == "31/12/2025"
    assert row8[4] == "31/12/2024"
    assert isinstance(row8[3], str)


def test_total_actif_present_dans_onglet(fm_ws, actif_gilac):
    """La ligne TOTAL ACTIF de l'onglet porte la valeur calculée."""
    _, ws = fm_ws
    for row in ws.iter_rows(min_row=10, values_only=True):
        if row and row[1] == "TOTAL ACTIF":
            assert abs(float(row[3]) - actif_gilac.total.valeur_n) <= 0.01
            break
    else:
        pytest.fail("Ligne TOTAL ACTIF introuvable dans l'onglet")
