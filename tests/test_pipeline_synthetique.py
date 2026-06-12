"""
Test end-to-end du pipeline sur données synthétiques (sans fichier client).

Remplace la couverture « données réelles » perdue avec la suppression des
fichiers GILAC : exécute main.run_pipeline en entier (FEC CP1252/TAB à
virgule française → contrôles → balance N vs N-1 → cycles → FM) et vérifie
les invariants comptables sur les livrables.
"""

from pathlib import Path

import openpyxl
import pytest

from main import run_pipeline
from tests.synthetic_data import (
    COMPTES_N1_SEULS,
    generer_balance_n1_xlsx,
    generer_fec_synthetique,
)

CLIENT = "DEMO"
DATE_CLOTURE = "31/12/2025"


# ---------------------------------------------------------------------------
# Fixtures (module : le pipeline ne tourne qu'une fois)
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def resultats(tmp_path_factory) -> dict:
    """Exécute le pipeline complet sur le FEC + balance N-1 synthétiques."""
    tmp = tmp_path_factory.mktemp("pipeline_synthetique")
    fec = generer_fec_synthetique(tmp / "DEMO_2025_FEC.txt")
    n1 = generer_balance_n1_xlsx(tmp / "balance_n1.xlsx")
    return run_pipeline(
        fec_path=str(fec),
        client=CLIENT,
        date_cloture=DATE_CLOTURE,
        n1_fm=str(n1),
        templates_dir=None,
        output_dir=str(tmp / "out"),
    )


@pytest.fixture(scope="module")
def wb_fm(resultats):
    return openpyxl.load_workbook(resultats["fm_path"], data_only=True)


# ---------------------------------------------------------------------------
# Contrôles du pipeline
# ---------------------------------------------------------------------------

def _controle(resultats, nom) -> tuple:
    par_nom = {c[0]: c for c in resultats["controles"]}
    assert nom in par_nom, f"Contrôle absent : {nom} ({sorted(par_nom)})"
    return par_nom[nom]


def test_controles_bloquants_ok(resultats):
    """Aucun contrôle BLOQUANT en échec (le pipeline est allé au bout)."""
    ko = [c for c in resultats["controles"] if c[3] == "BLOQUANT" and not c[1]]
    assert not ko


def test_controle_equilibre_bilan_ok(resultats):
    assert _controle(resultats, "Équilibre du bilan (AC-1)")[1]


def test_controle_coherence_resultat_ok(resultats):
    assert _controle(resultats, "Cohérence du résultat")[1]


def test_controle_coherence_treso_ok(resultats):
    """TN (FRNG − BFR) = trésorerie directe, malgré CBC, c/c associés
    et résultat en cours présents dans le FEC synthétique."""
    assert _controle(resultats, "Cohérence Tréso (TN vs trésorerie directe)")[1]


def test_controle_etats_detailles_ok(resultats):
    assert _controle(resultats, "Cohérence des états détaillés")[1]


def test_controle_pl_ok(resultats):
    assert _controle(resultats, "Cohérence du résultat net (P&L)")[1]


# ---------------------------------------------------------------------------
# Livrables
# ---------------------------------------------------------------------------

def test_livrables_generes(resultats):
    assert Path(resultats["fm_path"]).exists()
    assert Path(resultats["travail_path"]).exists()


def test_onglets_fm(wb_fm):
    attendus = ["Sommaire", "Balance N Vs N-1", "Bilan", "EBIT",
                "Actif détaillé", "Passif détaillé", "P&L détaillé",
                "Tréso", "AACE"]
    for nom in attendus:
        assert nom in wb_fm.sheetnames, f"Onglet manquant : {nom}"


def _lignes_balance(wb_fm):
    ws = wb_fm["Balance N Vs N-1"]
    lignes = {}
    for row in ws.iter_rows(min_row=10, max_col=5):
        if row[1].value is None:
            continue
        lignes[str(row[1].value)] = (row[3].value or 0.0, row[4].value or 0.0)
    return lignes


def test_balance_sommes_nulles(wb_fm):
    """Les soldes N et N-1 de la balance comparative rebouclent à zéro."""
    lignes = _lignes_balance(wb_fm)
    assert abs(sum(v[0] for v in lignes.values())) < 0.01
    assert abs(sum(v[1] for v in lignes.values())) < 0.01


def test_orphelins_n1_dans_balance(wb_fm):
    """Les comptes N-1 soldés en N figurent dans la balance (solde N = 0)."""
    lignes = _lignes_balance(wb_fm)
    for num in COMPTES_N1_SEULS:
        assert num in lignes, f"Compte N-1 soldé absent : {num}"
        solde_n, solde_n1 = lignes[num]
        assert solde_n == 0.0
        assert solde_n1 != 0.0


def test_treso_tn_egale_verification(wb_fm):
    """Dans l'onglet Tréso, la ligne d'écart de cohérence est ≈ 0."""
    ws = wb_fm["Tréso"]
    ecart = None
    for row in ws.iter_rows(max_col=6):
        label = next((c.value for c in row if isinstance(c.value, str)), None)
        if label and label.startswith("Écart de cohérence"):
            ecart = [c.value for c in row if isinstance(c.value, (int, float))]
    assert ecart is not None, "Ligne d'écart de cohérence absente de la Tréso"
    assert all(abs(v) <= 0.01 for v in ecart[:2]), f"Écart Tréso : {ecart}"


def test_bilan_equilibre_n_et_n1(wb_fm):
    """TOTAL ACTIF = TOTAL PASSIF en N et N-1 dans l'onglet Bilan."""
    ws = wb_fm["Bilan"]
    totaux = {}
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value in ("TOTAL ACTIF",
                                                        "TOTAL PASSIF"):
                vals = [x.value for x in row
                        if isinstance(x.value, (int, float))]
                totaux[c.value] = vals[:2]  # (N, N-1)
    assert set(totaux) == {"TOTAL ACTIF", "TOTAL PASSIF"}
    for idx in (0, 1):
        assert totaux["TOTAL ACTIF"][idx] == pytest.approx(
            totaux["TOTAL PASSIF"][idx], abs=0.01)
