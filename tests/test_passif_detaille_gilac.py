"""
Tests d'intégration de l'onglet Passif détaillé (cerfa 2051) — données GILAC.

La partition des structures détaillées reprend exactement les listes du
bilan synthétique : le total du Passif détaillé doit être strictement
égal au Total Passif du bilan (tolérance 1 K€ pour les arrondis).

Le FEC n'est parsé qu'une fois (fixtures module).
"""

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from src.engine.balance_builder import build
from src.engine.cycle_mapper import map_cycles
from src.engine.financial_engine import calculer_bilan, calculer_passif_detaille
from src.engine.liasse_fiscale_loader import load_liasse_fiscale
from src.parsers.fec_parser import parse
from src.parsers.mapping_parser import from_fm, from_pcg_config
from src.writers.fm_writer import write

DATA_DIR   = Path(__file__).parent.parent / "data"
CONFIG_DIR = Path(__file__).parent.parent / "src" / "config"

FEC_PATH = DATA_DIR / "GILAC_2025_12_31_FEC.txt"
FM_REF   = DATA_DIR / "FM GILAC.xlsx"
PCG_PATH = CONFIG_DIR / "mapping_pcg.yaml"

pytestmark = pytest.mark.skipif(
    not FEC_PATH.exists() or not FM_REF.exists(),
    reason="Fichiers GILAC absents (data/ non committé)",
)


# ---------------------------------------------------------------------------
# Fixtures (module : le FEC n'est parsé qu'une seule fois)
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def balance_mappee() -> pd.DataFrame:
    """Construit la balance mappée GILAC avec soldes N-1 (pipeline complet)."""
    df  = parse(FEC_PATH)
    mfm = from_fm(FM_REF)
    pcg = from_pcg_config(PCG_PATH)

    wb = openpyxl.load_workbook(FM_REF, read_only=True, data_only=True)
    ws = wb["Balance N Vs N-1"]
    balance_n1: dict = {}
    for row in ws.iter_rows(min_row=10, values_only=True):
        if row[1] is None:
            continue
        try:
            num = str(int(float(row[1])))
        except (ValueError, TypeError):
            continue
        solde = float(row[3]) if row[3] is not None else 0.0
        balance_n1[num] = {
            "libelle":  str(row[2]) if row[2] else "",
            "solde_ke": solde,
        }
    wb.close()

    bal = build(df, balance_n1)
    return map_cycles(bal, mfm, pcg)


@pytest.fixture(scope="module")
def liasse() -> dict:
    """Config liasse_fiscale complète (bilan + structures détaillées)."""
    return load_liasse_fiscale(from_pcg_config(PCG_PATH))


@pytest.fixture(scope="module")
def passif_gilac(balance_mappee, liasse):
    """Calcule le Passif détaillé GILAC."""
    return calculer_passif_detaille(balance_mappee, liasse)


@pytest.fixture(scope="module")
def bilan_gilac(balance_mappee, liasse):
    """Calcule le bilan synthétique GILAC (référence de cohérence)."""
    return calculer_bilan(balance_mappee, liasse)


@pytest.fixture(scope="module")
def fm_ws(balance_mappee, tmp_path_factory):
    """Génère le FM complet et retourne (sheetnames, onglet Passif détaillé)."""
    out = tmp_path_factory.mktemp("passif_detaille_gilac")
    pcg = from_pcg_config(PCG_PATH)
    fm_path = write(balance_mappee, "GILAC", "31/12/2025", out, pcg_config=pcg)
    wb = openpyxl.load_workbook(fm_path, read_only=True)
    yield wb.sheetnames, wb["Passif détaillé"]
    wb.close()


# ---------------------------------------------------------------------------
# Tests : cohérence avec le bilan synthétique
# ---------------------------------------------------------------------------

def test_total_passif_egal_bilan(passif_gilac, bilan_gilac):
    """Total Passif détaillé = Total Passif du bilan (N et N-1, ± 1 K€)."""
    assert abs(passif_gilac.total.valeur_n
               - bilan_gilac.total_passif.valeur_n) <= 1.0
    assert abs(passif_gilac.total.valeur_n1
               - bilan_gilac.total_passif.valeur_n1) <= 1.0


def test_reserves_partitionnent_le_poste_bilan(passif_gilac, bilan_gilac):
    """Réserve légale + statutaires + réglementées + autres = réserves bilan
    (1061 / 1063 / 1062+1064 / 1068 — convention cabinet pour 1062)."""
    postes_cp = {p.cle: p for s in passif_gilac.sections
                 for p in s.postes if s.cle == "capitaux_propres"}
    somme_reserves = (postes_cp["reserves_statutaires"].valeur_n
                      + postes_cp["reserves_reglementees"].valeur_n
                      + postes_cp["autres_reserves"].valeur_n)
    assert abs(postes_cp["reserve_legale"].valeur_n
               - bilan_gilac.postes["reserve_legale"].valeur_n) <= 0.01
    assert abs(somme_reserves
               - bilan_gilac.postes["autres_res"].valeur_n) <= 0.01


def test_capitaux_propres_coherents_bilan(passif_gilac, bilan_gilac):
    """Sous-total Capitaux propres = somme des postes équivalents du bilan."""
    sous_totaux = {s.cle: s.sous_total for s in passif_gilac.sections
                   if s.sous_total is not None}
    attendu = sum(
        bilan_gilac.postes[cle].valeur_n
        for cle in ("capital", "primes", "reserve_legale", "autres_res",
                    "report", "resultat", "resultat_encours", "subventions",
                    "prov_regl")
    )
    assert abs(sous_totaux["capitaux_propres"].valeur_n - attendu) <= 0.01


def test_total_egal_somme_des_postes(passif_gilac):
    """Le total est la somme de tous les postes de toutes les sections."""
    somme = sum(p.valeur_n for s in passif_gilac.sections for p in s.postes)
    assert abs(passif_gilac.total.valeur_n - somme) <= 0.01


# ---------------------------------------------------------------------------
# Tests : onglet Passif détaillé dans le FM généré
# ---------------------------------------------------------------------------

def test_onglet_existe(fm_ws):
    """L'onglet Passif détaillé existe dans le FM généré."""
    sheetnames, _ = fm_ws
    assert "Passif détaillé" in sheetnames


def test_onglet_apres_actif_detaille(fm_ws):
    """Le Passif détaillé suit immédiatement l'Actif détaillé."""
    sheetnames, _ = fm_ws
    assert (sheetnames.index("Passif détaillé")
            == sheetnames.index("Actif détaillé") + 1)


def test_total_passif_present_dans_onglet(fm_ws, passif_gilac):
    """La ligne TOTAL PASSIF de l'onglet porte la valeur calculée."""
    _, ws = fm_ws
    for row in ws.iter_rows(min_row=10, values_only=True):
        if row and row[1] == "TOTAL PASSIF":
            assert abs(float(row[3]) - passif_gilac.total.valeur_n) <= 0.01
            break
    else:
        pytest.fail("Ligne TOTAL PASSIF introuvable dans l'onglet")
