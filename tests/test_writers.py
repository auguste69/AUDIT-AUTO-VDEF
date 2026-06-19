"""
Tests pour src/writers/fm_writer.py et src/writers/styles.py.
Compare la sortie générée contre la structure du FM GILAC de référence.
"""

from pathlib import Path

import openpyxl
import pytest
import pandas as pd

from src.parsers.fec_parser import parse
from src.parsers.mapping_parser import from_fm, from_pcg_config
from src.engine.balance_builder import build
from src.engine.cycle_mapper import map_cycles
from src.writers.fm_writer import write

DATA_DIR   = Path(__file__).parent.parent / "data"
CONFIG_DIR = Path(__file__).parent.parent / "src" / "config"
OUTPUT_DIR = Path(__file__).parent.parent / "output" / "test"

FEC_PATH = DATA_DIR / "GILAC_2025_12_31_FEC.txt"
FM_REF   = DATA_DIR / "FM GILAC.xlsx"
PCG_PATH = CONFIG_DIR / "mapping_pcg.yaml"

pytestmark = pytest.mark.skipif(
    not FEC_PATH.exists() or not FM_REF.exists(),
    reason="Fichiers GILAC absents (données client retirées du dépôt — "
           "couverture assurée par tests/test_pipeline_synthetique.py)",
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def balance_mappee() -> pd.DataFrame:
    df  = parse(FEC_PATH)
    bal = build(df)
    mfm = from_fm(FM_REF)
    pcg = from_pcg_config(PCG_PATH)
    return map_cycles(bal, mfm, pcg)


@pytest.fixture(scope="module")
def fm_path(balance_mappee) -> Path:
    return write(balance_mappee, "GILAC", "31/12/2025", OUTPUT_DIR)


@pytest.fixture(scope="module")
def fm_wb(fm_path):
    return openpyxl.load_workbook(fm_path, read_only=True, data_only=True)


@pytest.fixture(scope="module")
def fm_ref_wb():
    return openpyxl.load_workbook(FM_REF, read_only=True, data_only=True)


# ---------------------------------------------------------------------------
# Tests : fichier généré
# ---------------------------------------------------------------------------

class TestFichierGenere:

    def test_fichier_existe(self, fm_path):
        assert fm_path.exists()

    def test_extension_xlsx(self, fm_path):
        assert fm_path.suffix == ".xlsx"

    def test_nom_fichier(self, fm_path):
        assert fm_path.name == "FM_GILAC_2025.xlsx"


# ---------------------------------------------------------------------------
# Tests : structure des onglets
# ---------------------------------------------------------------------------

class TestOnglets:

    ONGLETS_ATTENDUS = [
        "Sommaire", "Balance N Vs N-1", "Bilan", "EBIT",
        "Actif détaillé", "Passif détaillé", "P&L détaillé",
        "Tréso", "AACE",
        "C Propres0", "C PRC0", "F0", "I Incorp0", "I Corp0", "I Fi0",
        "S0", "A0", "V0", "P0", "E0", "T0", "X0",
    ]

    def test_nombre_onglets(self, fm_wb):
        assert len(fm_wb.sheetnames) == 22

    def test_noms_onglets(self, fm_wb):
        assert fm_wb.sheetnames == self.ONGLETS_ATTENDUS

    def test_ordre_cycles_canonique(self, fm_wb):
        """Les cycles suivent l'ordre PCG (C Propres en premier, X en dernier)."""
        _hors_cycles = {"Sommaire", "Balance N Vs N-1", "Bilan", "EBIT",
                        "Actif détaillé", "Passif détaillé", "P&L détaillé",
                        "Tréso", "AACE"}
        cycles = [s for s in fm_wb.sheetnames if s not in _hors_cycles]
        assert cycles[0]  == "C Propres0"
        assert cycles[-1] == "X0"


# ---------------------------------------------------------------------------
# Tests : onglet Balance N Vs N-1
# ---------------------------------------------------------------------------

class TestBalanceTab:

    def test_headers_row8(self, fm_wb):
        ws = fm_wb["Balance N Vs N-1"]
        rows = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))
        row = rows[0]
        assert row[0] == "Ref."
        assert row[5] == "Var. K€"
        assert row[6] == "Var. %"
        assert row[9] == "Cycle"
        assert row[10] == "EtatFi N"
        assert row[12] == "ComptaN"

    def test_titre_row3(self, fm_wb):
        ws = fm_wb["Balance N Vs N-1"]
        titre = ws.cell(row=3, column=4).value
        assert "31/12/2025" in str(titre)
        assert "31/12/2024" in str(titre)

    def test_client_row4(self, fm_wb):
        ws = fm_wb["Balance N Vs N-1"]
        assert ws.cell(row=4, column=4).value == "GILAC"

    def test_nombre_lignes_donnees(self, fm_wb):
        """270 comptes attendus à partir de row 10."""
        ws = fm_wb["Balance N Vs N-1"]
        lignes = [
            r for r in ws.iter_rows(min_row=10, values_only=True)
            if r[1] is not None and str(r[1]).strip() != ""
        ]
        assert len(lignes) == 270

    def test_comptes_tries_par_numero(self, fm_wb):
        ws = fm_wb["Balance N Vs N-1"]
        nums = [
            str(r[1]).strip()
            for r in ws.iter_rows(min_row=10, values_only=True)
            if r[1] is not None and str(r[1]).strip() != ""
        ]
        assert nums == sorted(nums)

    def test_solde_brut_pas_inverse(self, fm_wb, balance_mappee):
        """La Balance N Vs N-1 montre les valeurs brutes (non inversées)."""
        ws = fm_wb["Balance N Vs N-1"]
        # Compte 401000 (PASSIF, Solde_KE négatif car Créditeur)
        for row in ws.iter_rows(min_row=10, values_only=True):
            if row[1] == "401000":
                solde_attendu = float(
                    balance_mappee.loc[
                        balance_mappee["CompteNum"] == "401000", "Solde_KE"
                    ].iloc[0]
                )
                assert abs(float(row[3]) - solde_attendu) < 0.01
                break

    def test_cycle_colonne_j(self, fm_wb):
        """La colonne J (index 9) contient le cycle pour chaque compte."""
        ws = fm_wb["Balance N Vs N-1"]
        cycles_trouves = set()
        for row in ws.iter_rows(min_row=10, values_only=True):
            if row[1] is not None and row[9] is not None:
                cycles_trouves.add(str(row[9]).strip())
        assert len(cycles_trouves) == 13


# ---------------------------------------------------------------------------
# Tests : convention de signe dans les onglets cycle
# ---------------------------------------------------------------------------

class TestConventionSigne:
    """
    Règle : dans les onglets cycle,
      PASSIF   → valeur présentée = -Solde_KE
      PRODUITS → valeur présentée = -Solde_KE
      ACTIF    → valeur présentée = +Solde_KE
      CHARGES  → valeur présentée = +Solde_KE
    """

    def _valeur_cycle(self, fm_wb, onglet: str, compte_num: str):
        ws = fm_wb[onglet]
        for row in ws.iter_rows(min_row=10, values_only=True):
            if row[1] == compte_num:
                return row[3]  # colonne D = Solde N présenté
        return None

    def test_passif_signe_inverse(self, fm_wb, balance_mappee):
        """401000 (PASSIF, A0) doit être = -Solde_KE."""
        val_cycle = self._valeur_cycle(fm_wb, "A0", "401000")
        solde_ke  = float(
            balance_mappee.loc[balance_mappee["CompteNum"] == "401000", "Solde_KE"].iloc[0]
        )
        assert val_cycle is not None
        assert abs(float(val_cycle) - (-solde_ke)) < 0.01

    def test_charges_signe_identique(self, fm_wb, balance_mappee):
        """601100 (CHARGES, A0) doit être = +Solde_KE."""
        val_cycle = self._valeur_cycle(fm_wb, "A0", "601100")
        if val_cycle is None:
            pytest.skip("601100 absent de la balance 2025")
        solde_ke  = float(
            balance_mappee.loc[balance_mappee["CompteNum"] == "601100", "Solde_KE"].iloc[0]
        )
        assert abs(float(val_cycle) - solde_ke) < 0.01

    def test_actif_signe_identique(self, fm_wb, balance_mappee):
        """411000 (ACTIF, V0) doit être = +Solde_KE."""
        val_cycle = self._valeur_cycle(fm_wb, "V0", "411000")
        solde_ke  = float(
            balance_mappee.loc[balance_mappee["CompteNum"] == "411000", "Solde_KE"].iloc[0]
        )
        assert val_cycle is not None
        assert abs(float(val_cycle) - solde_ke) < 0.01

    def test_produits_signe_inverse(self, fm_wb, balance_mappee):
        """701000 (PRODUITS, V0) doit être = -Solde_KE."""
        val_cycle = self._valeur_cycle(fm_wb, "V0", "701000")
        solde_ke  = float(
            balance_mappee.loc[balance_mappee["CompteNum"] == "701000", "Solde_KE"].iloc[0]
        )
        assert val_cycle is not None
        assert abs(float(val_cycle) - (-solde_ke)) < 0.01

    def test_capital_passif_positif(self, fm_wb, balance_mappee):
        """101300 (PASSIF, C Propres0) doit être positif (capital = créditeur)."""
        val_cycle = self._valeur_cycle(fm_wb, "C Propres0", "101300")
        assert val_cycle is not None
        assert float(val_cycle) > 0


# ---------------------------------------------------------------------------
# Tests : structure des onglets cycle
# ---------------------------------------------------------------------------

class TestOngletsCycle:

    def test_titre_cycle(self, fm_wb):
        ws = fm_wb["A0"]
        titre = ws.cell(row=3, column=4).value
        assert "cycle A" in str(titre)

    def test_client_dans_cycle(self, fm_wb):
        ws = fm_wb["A0"]
        assert ws.cell(row=4, column=4).value == "GILAC"

    def test_headers_row8_cycle(self, fm_wb):
        ws = fm_wb["A0"]
        row = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]
        assert row[5] == "Var. K€"
        assert row[9] == "Cycle"

    def test_sections_presentes_A0(self, fm_wb):
        """A0 doit contenir ACTIF, PASSIF et CHARGES (PRODUITS absent dans A)."""
        ws = fm_wb["A0"]
        labels = {
            row[1] for row in ws.iter_rows(min_row=10, values_only=True)
            if row[1] in ("ACTIF", "PASSIF", "CHARGES", "PRODUITS")
        }
        assert "PASSIF"  in labels
        assert "CHARGES" in labels

    def test_sections_presentes_V0(self, fm_wb):
        """V0 doit contenir ACTIF, CHARGES et PRODUITS."""
        ws = fm_wb["V0"]
        labels = {
            row[1] for row in ws.iter_rows(min_row=10, values_only=True)
            if row[1] in ("ACTIF", "PASSIF", "CHARGES", "PRODUITS")
        }
        assert "ACTIF"   in labels
        assert "PRODUITS" in labels

    def test_nb_comptes_par_cycle(self, fm_wb, balance_mappee):
        """Chaque onglet contient exactement les comptes de son cycle."""
        for cycle in ["A", "V", "P", "E"]:
            nb_attendu = int((balance_mappee["cycle"] == cycle).sum())
            ws = fm_wb[f"{cycle}0"]
            nb_obtenus = sum(
                1 for row in ws.iter_rows(min_row=10, values_only=True)
                if row[1] is not None
                and row[1] not in ("ACTIF", "PASSIF", "CHARGES", "PRODUITS")
                and str(row[1]).strip() != ""
            )
            assert nb_obtenus == nb_attendu, (
                f"Cycle {cycle}: attendu {nb_attendu}, obtenu {nb_obtenus}"
            )

    def test_var_pct_na_pour_comptes_nouveaux(self, fm_wb, balance_mappee):
        """Comptes sans N-1 (Solde_N1_KE=0) → Var_PCT = 'n/a' dans l'onglet."""
        # 486001 est un nouveau compte (absent du FM N-1)
        ws = fm_wb["A0"]
        for row in ws.iter_rows(min_row=10, values_only=True):
            if row[1] == "486001":
                assert row[6] == "n/a", f"Var_PCT attendu 'n/a', obtenu {row[6]}"
                break

    def test_retour_sommaire_row1(self, fm_wb):
        """Row 1 de chaque onglet contient 'Retour sommaire'."""
        for onglet in ["A0", "V0", "Balance N Vs N-1"]:
            ws = fm_wb[onglet]
            val = ws.cell(row=1, column=1).value
            assert val == "Retour sommaire", f"Onglet {onglet}: row 1 = {val}"


# ---------------------------------------------------------------------------
# Tests : styles.py
# ---------------------------------------------------------------------------

class TestStyles:

    def test_import_styles(self):
        from src.writers.styles import (
            FONT_TITLE, FONT_SUBTITLE, FONT_HEADER, FONT_NORMAL, FONT_BOLD,
            FONT_SECTION, FONT_META, BORDER_BOTTOM_MED, BORDER_BOTTOM_HAIR,
            BORDER_TOP_THIN, NUM_KE, NUM_PCT,
        )
        assert FONT_TITLE.size == 16
        assert FONT_TITLE.bold is True
        assert FONT_HEADER.color.rgb.endswith("808080")
        assert NUM_KE == '#,##0;(#,##0);"-"'
        assert NUM_PCT == "0%"

    def test_remove_gridlines(self):
        from src.writers.styles import remove_gridlines
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        remove_gridlines(ws)
        assert ws.sheet_view.showGridLines is False

    def test_write_header_row(self):
        from src.writers.styles import write_header_row, BORDER_BOTTOM_MED
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        write_header_row(ws, 1, [(1, "Test"), (2, "Header")])
        assert ws.cell(1, 1).value == "Test"
        assert ws.cell(1, 2).value == "Header"
        # Bordure medium en dessous
        assert ws.cell(1, 1).border.bottom.style == "medium"
