"""
Tests d'équilibre du bilan synthétique pour le jeu de données GILAC.

Vérifie :
- Total Actif = Total Passif (écart < 0.5 K€)
- Aucune cellule "résiduel" dans l'onglet Bilan
"""
import pytest
import openpyxl
from pathlib import Path

DATA_DIR  = Path(__file__).parent.parent / "data"
FEC_PATH  = DATA_DIR / "GILAC_2025_12_31_FEC.txt"
N1_PATH   = DATA_DIR / "FM GILAC.xlsx"

pytestmark = pytest.mark.skipif(
    not FEC_PATH.exists() or not N1_PATH.exists(),
    reason="Fichiers GILAC absents (data/ non committé)",
)


@pytest.fixture(scope="module")
def fm_path(tmp_path_factory):
    """Exécute le pipeline et retourne le chemin du FM généré."""
    from main import run_pipeline
    out = tmp_path_factory.mktemp("bilan_fix")
    res = run_pipeline(
        fec_path=str(FEC_PATH),
        client="GILAC",
        date_cloture="31/12/2025",
        n1_fm=str(N1_PATH),
        output_dir=str(out),
    )
    return res["fm_path"]


@pytest.fixture(scope="module")
def ws_bilan(fm_path):
    wb = openpyxl.load_workbook(fm_path, read_only=True, data_only=True)
    ws = wb["Bilan"]
    return ws


def test_bilan_equilibre(ws_bilan):
    """Total Actif (col D) et Total Passif (col K) sur la ligne TOTAL doivent être égaux."""
    total_actif = total_passif = None
    for row in ws_bilan.iter_rows(values_only=True):
        label_a = row[1] if len(row) > 1  else None  # colonne B (actif)
        label_p = row[8] if len(row) > 8  else None  # colonne I (passif)
        if isinstance(label_a, str) and "TOTAL ACTIF" in label_a.upper():
            total_actif  = row[3]   # colonne D
        if isinstance(label_p, str) and "TOTAL PASSIF" in label_p.upper():
            total_passif = row[10]  # colonne K
    assert total_actif  is not None, "Ligne 'TOTAL ACTIF' introuvable dans l'onglet Bilan"
    assert total_passif is not None, "Ligne 'TOTAL PASSIF' introuvable dans l'onglet Bilan"
    assert abs(float(total_actif) - float(total_passif)) < 0.5, (
        f"Bilan déséquilibré : Actif={total_actif} K€, Passif={total_passif} K€, "
        f"écart={abs(float(total_actif) - float(total_passif)):.3f} K€"
    )


def test_aucun_residuel(ws_bilan):
    """Aucune cellule de l'onglet Bilan ne doit contenir le mot 'résiduel'."""
    for row in ws_bilan.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str) and "résiduel" in cell.lower():
                pytest.fail(f"Cellule 'résiduel' trouvée dans l'onglet Bilan : {cell!r}")
