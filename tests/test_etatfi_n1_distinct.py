"""
Tests des colonnes distinctes EtatFi N / EtatFi N-1 et ComptaN / ComptaN-1.

Vérifie la correction de la perte d'information sur les reclassements
inter-exercices :
  1. mapping_parser.from_fm lit les 4 colonnes K/L/M/N du FM
  2. cycle_mapper.map_cycles propage les 4 champs dans la balance mappée
  3. fm_writer écrit 4 colonnes distinctes (K=EtatFi N, L=EtatFi N-1,
     M=ComptaN, N=ComptaN-1)
  4. les aliases historiques etatfi / compta sont conservés partout
"""

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from src.engine.cycle_mapper import map_cycles
from src.parsers.mapping_parser import from_fm, from_pcg_config
from src.writers.fm_writer import write

DATA_DIR   = Path(__file__).parent.parent / "data"
CONFIG_DIR = Path(__file__).parent.parent / "src" / "config"

FM_REF   = DATA_DIR / "FM GILAC.xlsx"
PCG_PATH = CONFIG_DIR / "mapping_pcg.yaml"

# Comptes GILAC connus pour être reclassés entre N-1 et N (cf. FM source)
_COMPTES_RECLASSES_ATTENDUS = ["413000", "421000", "425000"]

_skip_si_donnees_absentes = pytest.mark.skipif(
    not FM_REF.exists(),
    reason="Fichier FM GILAC absent (data/ non committé)",
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def pcg() -> dict:
    """Config PCG réelle du cabinet."""
    return from_pcg_config(PCG_PATH)


@pytest.fixture(scope="module")
def mapping_gilac() -> dict:
    """Mapping extrait du FM GILAC réel."""
    if not FM_REF.exists():
        pytest.skip("Fichier FM GILAC absent")
    return from_fm(FM_REF)


@pytest.fixture()
def mapping_synthetique() -> dict:
    """Mapping de type from_fm avec un compte reclassé entre N-1 et N."""
    return {
        # Compte reclassé : Actif en N, Passif en N-1
        "413000": {
            "cycle":     "V",
            "etatfi_n":  "Clients et comptes rattachés",
            "etatfi_n1": "Clients créditeurs",
            "compta_n":  "Actif",
            "compta_n1": "Passif",
            "etatfi":    "Clients et comptes rattachés",
            "compta":    "Actif",
            "ref":       "V0",
        },
        # Compte stable : mêmes valeurs N et N-1
        "601000": {
            "cycle":     "A",
            "etatfi_n":  "Achats matières premières",
            "etatfi_n1": "Achats matières premières",
            "compta_n":  "Charges",
            "compta_n1": "Charges",
            "etatfi":    "Achats matières premières",
            "compta":    "Charges",
            "ref":       "A0",
        },
    }


@pytest.fixture()
def balance_synthetique() -> pd.DataFrame:
    """Balance minimale (structure balance_builder.build) avec 3 comptes."""
    return pd.DataFrame({
        "CompteNum":   ["413000", "601000", "411500"],
        "CompteLib":   ["Clients - effets", "Achats MP", "Clients divers"],
        "Debit":       [10000.0, 5000.0, 2000.0],
        "Credit":      [0.0, 0.0, 0.0],
        "Solde":       [10000.0, 5000.0, 2000.0],
        "Solde_KE":    [10.0, 5.0, 2.0],
        "Solde_N1_KE": [-8.0, 4.0, 0.0],
        "Var_KE":      [18.0, 1.0, 2.0],
        "Var_PCT":     [2.25, 0.25, "n/a"],
    })


# ---------------------------------------------------------------------------
# 1. Propagation sur mapping synthétique
# ---------------------------------------------------------------------------

class TestPropagationSynthetique:
    def test_quatre_champs_propages(self, balance_synthetique,
                                    mapping_synthetique, pcg):
        """etatfi_n != etatfi_n1 du mapping arrive dans la balance mappée."""
        bm = map_cycles(balance_synthetique, mapping_synthetique, pcg)
        r = bm[bm["CompteNum"] == "413000"].iloc[0]

        assert r["etatfi_n"]  == "Clients et comptes rattachés"
        assert r["etatfi_n1"] == "Clients créditeurs"
        assert r["compta_n"]  == "Actif"
        assert r["compta_n1"] == "Passif"
        assert r["etatfi_n"] != r["etatfi_n1"]
        assert r["compta_n"] != r["compta_n1"]

    def test_compte_stable_valeurs_identiques(self, balance_synthetique,
                                              mapping_synthetique, pcg):
        """Un compte non reclassé garde les mêmes valeurs N et N-1."""
        bm = map_cycles(balance_synthetique, mapping_synthetique, pcg)
        r = bm[bm["CompteNum"] == "601000"].iloc[0]

        assert r["etatfi_n"] == r["etatfi_n1"] == "Achats matières premières"
        assert r["compta_n"] == r["compta_n1"] == "Charges"

    def test_nouveau_compte_fallback_pcg_n1_vide(self, balance_synthetique,
                                                 mapping_synthetique, pcg):
        """Compte absent du FM (411500) : valeurs N-1 vides (pas d'historique)."""
        bm = map_cycles(balance_synthetique, mapping_synthetique, pcg)
        r = bm[bm["CompteNum"] == "411500"].iloc[0]

        assert r["compta_n"] == "Actif"   # fallback PCG (411 = clients)
        assert r["compta_n1"] == ""
        assert r["etatfi_n1"] == ""

    def test_balance_source_non_modifiee(self, balance_synthetique,
                                         mapping_synthetique, pcg):
        """map_cycles ne modifie pas le DataFrame d'entrée."""
        colonnes_avant = list(balance_synthetique.columns)
        map_cycles(balance_synthetique, mapping_synthetique, pcg)
        assert list(balance_synthetique.columns) == colonnes_avant
        assert "etatfi_n1" not in balance_synthetique.columns


# ---------------------------------------------------------------------------
# 2. Aliases historiques conservés
# ---------------------------------------------------------------------------

class TestAliases:
    def test_aliases_balance_mappee(self, balance_synthetique,
                                    mapping_synthetique, pcg):
        """Les colonnes etatfi et compta existent et valent les valeurs N."""
        bm = map_cycles(balance_synthetique, mapping_synthetique, pcg)

        assert "etatfi" in bm.columns
        assert "compta" in bm.columns
        assert (bm["etatfi"] == bm["etatfi_n"]).all()
        assert (bm["compta"] == bm["compta_n"]).all()

    @_skip_si_donnees_absentes
    def test_aliases_mapping_from_fm(self, mapping_gilac):
        """from_fm conserve les clés etatfi / compta (aliases des valeurs N)."""
        for info in mapping_gilac.values():
            assert "etatfi" in info
            assert "compta" in info
            assert info["etatfi"] == info["etatfi_n"]
            assert info["compta"] == info["compta_n"]

    def test_mapping_legacy_sans_cles_n1(self, balance_synthetique, pcg):
        """Un mapping ancien (clés etatfi/compta seules) reste accepté."""
        mapping_legacy = {
            "413000": {"cycle": "V", "etatfi": "Clients", "compta": "Actif",
                       "ref": "V0"},
        }
        bm = map_cycles(balance_synthetique, mapping_legacy, pcg)
        r = bm[bm["CompteNum"] == "413000"].iloc[0]
        assert r["etatfi_n"] == "Clients"
        assert r["compta_n"] == "Actif"
        assert r["etatfi_n1"] == ""
        assert r["compta_n1"] == ""


# ---------------------------------------------------------------------------
# 3. Intégration — données GILAC réelles
# ---------------------------------------------------------------------------

@_skip_si_donnees_absentes
class TestIntegrationGilac:
    def test_cles_n_et_n1_presentes(self, mapping_gilac):
        """Chaque entrée du mapping a les 4 clés distinctes."""
        for info in mapping_gilac.values():
            for cle in ("etatfi_n", "etatfi_n1", "compta_n", "compta_n1"):
                assert cle in info

    def test_comptes_reclasses_etatfi(self, mapping_gilac):
        """Au moins 1 compte avec etatfi_n != etatfi_n1 (22 attendus)."""
        reclasses = [
            num for num, info in mapping_gilac.items()
            if info["etatfi_n"] != info["etatfi_n1"]
        ]
        assert len(reclasses) >= 1, "Aucun reclassement EtatFi détecté"
        for compte in _COMPTES_RECLASSES_ATTENDUS:
            assert compte in reclasses, (
                f"Compte {compte} attendu parmi les reclassés EtatFi"
            )

    def test_comptes_reclasses_compta(self, mapping_gilac):
        """Au moins 1 compte avec compta_n != compta_n1."""
        reclasses = [
            num for num, info in mapping_gilac.items()
            if info["compta_n"] != info["compta_n1"]
        ]
        assert len(reclasses) >= 1, "Aucun reclassement Compta détecté"
        assert "413000" in reclasses

    def test_compte_413000_valeurs(self, mapping_gilac):
        """Le compte 413000 est Actif en N, Passif en N-1 (FM source)."""
        info = mapping_gilac["413000"]
        assert info["compta_n"]  == "Actif"
        assert info["compta_n1"] == "Passif"
        assert info["etatfi_n"] != info["etatfi_n1"]


# ---------------------------------------------------------------------------
# 4. FM généré — colonnes K/L distinctes
# ---------------------------------------------------------------------------

class TestFmGenere:
    def test_colonnes_k_l_m_n_distinctes(self, balance_synthetique,
                                         mapping_synthetique, pcg, tmp_path):
        """Le FM écrit des valeurs distinctes en K/L et M/N pour un reclassé."""
        bm = map_cycles(balance_synthetique, mapping_synthetique, pcg)
        chemin = write(bm, "TESTCLIENT", "31/12/2025", tmp_path,
                       pcg_config=pcg)

        wb = openpyxl.load_workbook(chemin)
        ws = wb["Balance N Vs N-1"]

        ligne_413000 = None
        for row in ws.iter_rows(min_row=10, max_row=ws.max_row):
            if str(row[1].value) == "413000":  # colonne B
                ligne_413000 = row
                break
        assert ligne_413000 is not None, "Compte 413000 absent du FM généré"

        etatfi_n  = ligne_413000[10].value  # colonne K
        etatfi_n1 = ligne_413000[11].value  # colonne L
        compta_n  = ligne_413000[12].value  # colonne M
        compta_n1 = ligne_413000[13].value  # colonne N

        assert etatfi_n  == "Clients et comptes rattachés"
        assert etatfi_n1 == "Clients créditeurs"
        assert compta_n  == "Actif"
        assert compta_n1 == "Passif"
        assert etatfi_n != etatfi_n1
        assert compta_n != compta_n1
        wb.close()

    def test_headers_balance(self, balance_synthetique, mapping_synthetique,
                             pcg, tmp_path):
        """Les headers K/L/M/N de la balance sont distincts."""
        bm = map_cycles(balance_synthetique, mapping_synthetique, pcg)
        chemin = write(bm, "TESTCLIENT", "31/12/2025", tmp_path,
                       pcg_config=pcg)

        wb = openpyxl.load_workbook(chemin)
        ws = wb["Balance N Vs N-1"]
        assert ws.cell(row=8, column=11).value == "EtatFi N"
        assert ws.cell(row=8, column=12).value == "EtatFi N-1"
        assert ws.cell(row=8, column=13).value == "ComptaN"
        assert ws.cell(row=8, column=14).value == "ComptaN-1"
        wb.close()
