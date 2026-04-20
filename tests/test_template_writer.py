"""
Tests pour src/writers/template_writer.py.
Vérifie le renommage, les remplacements de placeholders, et la structure du ZIP.
"""

import io
import zipfile
from pathlib import Path

import openpyxl
import pytest

from src.parsers.mapping_parser import from_pcg_config
from src.writers.template_writer import write, _detecter_cycle, _normaliser

DATA_DIR      = Path(__file__).parent.parent / "data"
TEMPLATES_DIR = DATA_DIR / "templates"
CONFIG_DIR    = Path(__file__).parent.parent / "src" / "config"
OUTPUT_DIR    = Path(__file__).parent.parent / "output" / "test_templates"

CLIENT       = "GILAC"
DATE_CLOTURE = "31/12/2025"


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def pcg():
    return from_pcg_config(CONFIG_DIR / "mapping_pcg.yaml")


@pytest.fixture(scope="module")
def zip_path(pcg) -> Path:
    return write(TEMPLATES_DIR, CLIENT, DATE_CLOTURE, OUTPUT_DIR, pcg["templates"])


@pytest.fixture(scope="module")
def zip_contents(zip_path) -> dict:
    """Retourne {nom_fichier: bytes} pour chaque fichier du ZIP."""
    with zipfile.ZipFile(zip_path) as zf:
        return {name: zf.read(name) for name in zf.namelist()}


def _open_from_zip(zip_contents: dict, nom: str):
    """Ouvre un workbook openpyxl depuis les bytes du ZIP."""
    data = zip_contents[nom]
    return openpyxl.load_workbook(io.BytesIO(data))


# ---------------------------------------------------------------------------
# Tests : fichier ZIP
# ---------------------------------------------------------------------------

class TestZip:

    def test_zip_existe(self, zip_path):
        assert zip_path.exists()

    def test_nom_zip(self, zip_path):
        assert zip_path.name == "FT_GILAC_2025.zip"

    def test_13_fichiers(self, zip_contents):
        assert len(zip_contents) == 13

    def test_tous_les_cycles_presents(self, zip_contents):
        """Chacun des 13 cycles a son fichier dans le ZIP."""
        noms = set(zip_contents.keys())
        cycles_attendus = ["A", "C", "E", "F", "I", "P", "S", "T", "V", "X"]
        for cycle in cycles_attendus:
            trouve = any(f"_{cycle}" in n or f"_{cycle.lower()}" in n.lower()
                        for n in noms)
            assert trouve, f"Aucun fichier pour le cycle {cycle}"

    def test_prefixe_annee_client(self, zip_contents):
        """Tous les fichiers commencent par '2025_12_GILAC_'."""
        for nom in zip_contents:
            assert nom.startswith("2025_12_GILAC_"), f"Préfixe incorrect : {nom}"

    def test_plus_de_prefixe_20XX(self, zip_contents):
        """Aucun fichier ne doit garder le préfixe 20XX_XX_YYY_."""
        for nom in zip_contents:
            assert "20XX" not in nom, f"Préfixe 20XX non remplacé : {nom}"

    def test_tous_xlsx(self, zip_contents):
        for nom in zip_contents:
            assert nom.endswith(".xlsx"), f"Extension incorrecte : {nom}"


# ---------------------------------------------------------------------------
# Tests : remplacement des placeholders
# ---------------------------------------------------------------------------

class TestPlaceholders:

    def _cellules_contenant(self, wb, texte: str) -> list:
        """Retourne [(sheet, row, col, valeur)] pour toutes cellules contenant texte."""
        resultats = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and texte in cell.value:
                        resultats.append((sheet_name, cell.row, cell.column, cell.value))
        return resultats

    def test_nomclient_remplace_dans_A(self, zip_contents):
        wb = _open_from_zip(zip_contents, "2025_12_GILAC_A_Fournisseurs-Achats.xlsx")
        assert not self._cellules_contenant(wb, "#NomClient"), \
            "#NomClient encore présent dans A Fournisseurs"

    def test_nomclient_valeur_correcte_dans_A(self, zip_contents):
        wb = _open_from_zip(zip_contents, "2025_12_GILAC_A_Fournisseurs-Achats.xlsx")
        # R1C4 doit contenir le client
        val = wb["A Synthèse"].cell(row=1, column=4).value
        assert val == CLIENT, f"Attendu '{CLIENT}', obtenu '{val}'"

    def test_datecloture_remplace_dans_A(self, zip_contents):
        wb = _open_from_zip(zip_contents, "2025_12_GILAC_A_Fournisseurs-Achats.xlsx")
        assert not self._cellules_contenant(wb, "#DateClôture"), \
            "#DateClôture encore présent dans A Fournisseurs"

    def test_datecloture_valeur_correcte(self, zip_contents):
        wb = _open_from_zip(zip_contents, "2025_12_GILAC_A_Fournisseurs-Achats.xlsx")
        # R7C13 = date de clôture
        val = wb["A Synthèse"].cell(row=7, column=13).value
        assert val == "31/12/2025", f"Attendu '31/12/2025', obtenu '{val}'"

    def test_placeholder_date_remplace_dans_C_capitaux(self, zip_contents):
        """Le template C Capitaux a des '#Date' dans C100 et C300."""
        wb = _open_from_zip(zip_contents, "2025_12_GILAC_C_Capitaux propres.xlsx")
        assert not self._cellules_contenant(wb, "#Date"), \
            "#Date encore présent dans C Capitaux propres"

    def test_tous_templates_sans_nomclient(self, zip_contents):
        """Aucun fichier du ZIP ne doit contenir '#NomClient'."""
        for nom, data in zip_contents.items():
            wb = openpyxl.load_workbook(io.BytesIO(data))
            restants = self._cellules_contenant(wb, "#NomClient")
            assert not restants, f"{nom} contient encore #NomClient : {restants}"

    def test_tous_templates_sans_datecloture(self, zip_contents):
        """Aucun fichier du ZIP ne doit contenir '#DateClôture'."""
        for nom, data in zip_contents.items():
            wb = openpyxl.load_workbook(io.BytesIO(data))
            restants = self._cellules_contenant(wb, "#DateClôture")
            assert not restants, f"{nom} contient encore #DateClôture : {restants}"

    def test_placeholders_texte_intacts(self, zip_contents):
        """
        Les '#' dans les instructions de travail (ex: '#411', '#512') ne doivent
        pas être modifiés — seuls les placeholders exacts sont remplacés.
        """
        wb = _open_from_zip(zip_contents, "2025_12_GILAC_V_Clients-Ventes.xlsx")
        # Vérifier qu'une référence de compte genre '#411' existe encore (texte d'instruction)
        tous_textes = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "#" in cell.value:
                        tous_textes.append(cell.value)
        # Il doit rester des # dans les textes d'instruction (comptes, formules)
        assert len(tous_textes) > 0, "Tous les # ont été supprimés (trop agressif)"


# ---------------------------------------------------------------------------
# Tests : structure des feuilles
# ---------------------------------------------------------------------------

class TestStructure:

    def test_nb_feuilles_A(self, zip_contents):
        wb = _open_from_zip(zip_contents, "2025_12_GILAC_A_Fournisseurs-Achats.xlsx")
        assert len(wb.sheetnames) == 14

    def test_nb_feuilles_V(self, zip_contents):
        wb = _open_from_zip(zip_contents, "2025_12_GILAC_V_Clients-Ventes.xlsx")
        assert len(wb.sheetnames) == 18

    def test_nb_feuilles_S(self, zip_contents):
        wb = _open_from_zip(zip_contents, "2025_12_GILAC_S_Stocks.xlsx")
        assert len(wb.sheetnames) == 13

    def test_feuille_synth_presente_dans_chaque_template(self, zip_contents):
        for nom, data in zip_contents.items():
            wb = openpyxl.load_workbook(io.BytesIO(data))
            a_synth = any("Synthèse" in s or "synth" in s.lower() for s in wb.sheetnames)
            assert a_synth, f"{nom} n'a pas de feuille Synthèse"

    def test_gridlines_supprimes(self, zip_contents):
        """Toutes les feuilles de tous les fichiers n'ont plus de quadrillage."""
        for nom, data in zip_contents.items():
            wb = openpyxl.load_workbook(io.BytesIO(data))
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                assert ws.sheet_view.showGridLines is False, \
                    f"Quadrillage encore actif : {nom}::{sheet_name}"


# ---------------------------------------------------------------------------
# Tests : _detecter_cycle et _normaliser
# ---------------------------------------------------------------------------

class TestDetecterCycle:

    def test_normaliser_accents(self):
        assert _normaliser("Résultat") == "resultat"

    def test_normaliser_espaces_underscores(self):
        assert _normaliser("I_Immobilisations corp") == _normaliser("I_Immobilisations_corp")

    def test_normaliser_cle_x(self):
        # X_Re_sultat et Résultat doivent matcher
        assert "xresultat" in _normaliser("X_Résultat exceptionnel")
        assert _normaliser("X_Re_sultat") == "xresultat"

    def test_detecter_cycle_A(self, pcg):
        cycle = _detecter_cycle(
            "20XX_XX_YYY_A_Fournisseurs-Achats.xlsx", pcg["templates"]
        )
        assert cycle == "A"

    def test_detecter_cycle_I_corp(self, pcg):
        cycle = _detecter_cycle(
            "20XX_XX_YYY_I_Immobilisations corp.xlsx", pcg["templates"]
        )
        assert cycle == "I Corp"

    def test_detecter_cycle_I_fi(self, pcg):
        cycle = _detecter_cycle(
            "20XX_XX_YYY_I_Immobilisations financières.xlsx", pcg["templates"]
        )
        assert cycle == "I Fi"

    def test_detecter_cycle_X(self, pcg):
        cycle = _detecter_cycle(
            "20XX_XX_YYY_X_Résultat exceptionnel.xlsx", pcg["templates"]
        )
        assert cycle == "X"

    def test_detecter_cycle_inconnu(self, pcg):
        assert _detecter_cycle("fichier_inconnu.xlsx", pcg["templates"]) is None


# ---------------------------------------------------------------------------
# Tests : erreurs
# ---------------------------------------------------------------------------

class TestErreurs:

    def test_dossier_inexistant(self, pcg):
        with pytest.raises(FileNotFoundError):
            write("/tmp/dossier_inexistant", CLIENT, DATE_CLOTURE, OUTPUT_DIR)

    def test_dossier_vide(self, tmp_path, pcg):
        with pytest.raises(ValueError, match="Aucun template"):
            write(tmp_path, CLIENT, DATE_CLOTURE, OUTPUT_DIR)

    def test_sans_mapping_traite_tout(self, tmp_path):
        """Sans mapping, tous les .xlsx sont traités (pas de filtre cycle)."""
        import shutil
        src = TEMPLATES_DIR / "20XX_XX_YYY_A_Fournisseurs-Achats.xlsx"
        shutil.copy(src, tmp_path / src.name)
        zip_p = write(tmp_path, "TEST", "31/12/2025", tmp_path / "out", None)
        with zipfile.ZipFile(zip_p) as zf:
            assert len(zf.namelist()) == 1
