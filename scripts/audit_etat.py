#!/usr/bin/env python3
"""
audit_etat.py — Audit non-destructif de l'état du projet audit-automation-2.

Génère output/AUDIT_ETAT.md avec 5 sections :
  1. Fichiers attendus
  2. Livrables
  3. Dette technique
  4. Cibles chiffrées GILAC
  5. Tests

Idempotent : relançable sans side-effect.
"""

import io
import logging
import re
import subprocess
import sys
import zipfile
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional

# ---------------------------------------------------------------------------
# Chemins
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Handler de capture de logs
# ---------------------------------------------------------------------------
class _CapturingHandler(logging.Handler):
    """Capture les messages de log dans une liste (sans les afficher)."""

    def __init__(self, level: int = logging.WARNING) -> None:
        super().__init__(level)
        self.records: List[str] = []
        self.setFormatter(logging.Formatter("%(levelname)s %(name)s: %(message)s"))

    def emit(self, record: logging.LogRecord) -> None:
        self.records.append(self.format(record))


# ---------------------------------------------------------------------------
# Section 1 — Fichiers attendus
# ---------------------------------------------------------------------------
def _section1() -> List[str]:
    lignes = ["## 1. Fichiers attendus\n"]
    attendus = [
        (ROOT / "src/engine/financial_engine.py",      "fichier"),
        (ROOT / "src/engine/account_matcher.py",       "fichier"),
        (ROOT / "src/engine/liasse_fiscale_loader.py", "fichier"),
        (ROOT / "src/models",                          "dossier"),
        (ROOT / "src/writers/fm",                      "dossier"),
        (ROOT / "src/writers/worksheet_copy.py",       "fichier"),
        (ROOT / "src/parsers/balance_n1_loader.py",    "fichier"),
        (ROOT / "tests/fixtures",                      "dossier"),
    ]
    for path, type_ in attendus:
        present = path.exists() if type_ == "fichier" else path.is_dir()
        statut = "✓ présent" if present else "✗ absent"
        lignes.append(f"- `{path.relative_to(ROOT)}` ({type_}) : {statut}")
    return lignes


# ---------------------------------------------------------------------------
# Section 2 — Livrables
# ---------------------------------------------------------------------------
def _section2(
    resultats: Optional[Dict],
    pipeline_error: Optional[Exception],
    output_tmp: Path,
) -> List[str]:
    lignes = ["## 2. Livrables\n"]

    cible_onglets = [
        "Sommaire", "Balance N Vs N-1", "Bilan", "EBIT", "Actif détaillé",
        "Passif détaillé", "P&L détaillé", "Tréso", "AACE",
        "C Propres0", "C PRC0", "F0", "I Incorp0", "I Corp0", "I Fi0",
        "S0", "A0", "V0", "P0", "E0", "T0", "X0",
    ]

    # 2a. Pipeline
    lignes.append("### 2a. Pipeline GILAC\n")
    if pipeline_error is not None:
        lignes.append(f"- Pipeline : ✗ ERREUR — {pipeline_error}")
        return lignes
    if resultats is None:
        lignes.append("- Pipeline : non exécuté (FEC introuvable)")
        return lignes

    lignes.append(
        f"- Pipeline : ✓ terminé"
        f" ({resultats['fec_lignes']} écritures,"
        f" {resultats['nb_comptes']} comptes,"
        f" {len(resultats['controles'])} contrôles)"
    )

    # 2b. Onglets du FM
    lignes.append("\n### 2b. Onglets du FM produit\n")
    fm_path = output_tmp / "FM_GILAC_2025.xlsx"
    if not fm_path.exists():
        logger.error("FM introuvable : %s", fm_path)
        lignes.append("- FM : ✗ introuvable")
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(fm_path), read_only=True)
            onglets = list(wb.sheetnames)
            wb.close()
            lignes.append(f"Onglets produits ({len(onglets)}) : {', '.join(onglets)}\n")
            manquants  = [o for o in cible_onglets if o not in onglets]
            non_prevus = [o for o in onglets      if o not in cible_onglets]
            lignes.append(
                f"- Manquants : {', '.join(manquants) if manquants else 'aucun'}"
            )
            lignes.append(
                f"- Présents non prévus : "
                f"{', '.join(non_prevus) if non_prevus else 'aucun'}"
            )
        except Exception as exc:
            logger.error("Lecture FM : %s", exc)
            lignes.append(f"- FM : ✗ erreur lecture — {exc}")

    # 2c. ZIP templates
    lignes.append("\n### 2c. Contenu du ZIP templates\n")
    zip_path = output_tmp / "FT_GILAC_2025.zip"
    if not zip_path.exists():
        logger.warning("ZIP introuvable : %s", zip_path)
        lignes.append("- ZIP : ✗ introuvable")
    else:
        onglets_fm_attendus = {
            "AACE", "A0", "V0", "Tréso", "F0",
            "C Propres0", "C PRC0", "E0", "P0", "S0", "T0", "X0",
            "I Corp0", "I Incorp0", "I Fi0",
        }
        try:
            import openpyxl
            with zipfile.ZipFile(str(zip_path), "r") as zf:
                noms = sorted(zf.namelist())
                lignes.append(f"Fichiers dans le ZIP ({len(noms)}) :\n")
                for nom in noms:
                    try:
                        with zf.open(nom) as f:
                            wb_t = openpyxl.load_workbook(
                                io.BytesIO(f.read()), read_only=True
                            )
                            ong_t = list(wb_t.sheetnames)
                            wb_t.close()
                        fm_trouves = [o for o in ong_t if o in onglets_fm_attendus]
                        p1_label = "✓ P1 implémenté" if fm_trouves else "✗ P1 absent"
                        lignes.append(
                            f"- `{nom}` : {len(ong_t)} onglets —"
                            f" FM trouvés : {fm_trouves or 'aucun'} — {p1_label}"
                        )
                    except Exception as exc_t:
                        logger.error("Lecture template %s : %s", nom, exc_t)
                        lignes.append(f"- `{nom}` : ✗ erreur — {exc_t}")
        except Exception as exc_zip:
            logger.error("Lecture ZIP : %s", exc_zip)
            lignes.append(f"- ZIP : ✗ erreur — {exc_zip}")

    return lignes


# ---------------------------------------------------------------------------
# Section 3 — Dette technique
# ---------------------------------------------------------------------------
def _section3() -> List[str]:
    lignes = ["## 3. Dette technique\n"]

    # 3a. Préfixes PCG en dur dans src/
    lignes.append('### 3a. Préfixes PCG en dur dans `src/` (pattern `["NNN"]`)\n')
    pattern_pcg = re.compile(r'\[\s*"(\d{2,4})"')
    matches: List[str] = []
    for py in sorted((ROOT / "src").rglob("*.py")):
        try:
            for i, line in enumerate(
                py.read_text(encoding="utf-8", errors="replace").splitlines(), 1
            ):
                if pattern_pcg.search(line):
                    extrait = line.strip()[:80]
                    matches.append(
                        f"- `{py.relative_to(ROOT)}:{i}` → `{extrait}`"
                    )
        except Exception as exc:
            logger.error("Lecture %s : %s", py, exc)
    if matches:
        lignes.extend(matches)
    else:
        lignes.append("Aucun préfixe PCG en dur détecté.")

    # 3b. Section liasse_fiscale dans le YAML
    lignes.append("\n### 3b. Section `liasse_fiscale` dans `mapping_pcg.yaml`\n")
    yaml_path = ROOT / "src/config/mapping_pcg.yaml"
    if yaml_path.exists():
        text = yaml_path.read_text(encoding="utf-8")
        if "liasse_fiscale" in text:
            sous = re.findall(r"^  (\w[\w_]+)\s*:", text, re.MULTILINE)
            lignes.append(
                f"✓ Présente — sous-sections trouvées : "
                f"{', '.join(sous) if sous else 'aucune'}"
            )
        else:
            lignes.append("✗ Absente")
    else:
        lignes.append(
            f"✗ Fichier YAML introuvable : {yaml_path.relative_to(ROOT)}"
        )

    # 3c. Fonctions clés dans template_writer.py
    lignes.append("\n### 3c. Fonctions / méthodes clés dans `template_writer.py`\n")
    tw = ROOT / "src/writers/template_writer.py"
    if tw.exists():
        txt = tw.read_text(encoding="utf-8")
        for name in ["_injecter_balance_cycle", "move_sheet", "_inserer_onglet_fm"]:
            present = name in txt
            lignes.append(f"- `{name}` : {'✓ présent' if present else '✗ absent'}")
    else:
        lignes.append("✗ `template_writer.py` introuvable")

    # 3d. Duplication extraction N-1 (iter_rows(min_row=10)
    lignes.append(
        "\n### 3d. Duplication extraction N-1 (`iter_rows(min_row=10`)\n"
    )
    for nom in ["main.py", "app.py"]:
        p = ROOT / nom
        if p.exists():
            n = p.read_text(encoding="utf-8", errors="replace").count(
                "iter_rows(min_row=10"
            )
            lignes.append(f"- `{nom}` : {n} occurrence(s)")
        else:
            lignes.append(f"- `{nom}` : ✗ introuvable")

    return lignes


# ---------------------------------------------------------------------------
# Section 4 — Cibles chiffrées GILAC
# ---------------------------------------------------------------------------
def _section4(warnings_captures: List[str]) -> List[str]:
    lignes = ["## 4. Cibles chiffrées GILAC\n"]

    fec_path = ROOT / "data/GILAC_2025_12_31_FEC.txt"
    if not fec_path.exists():
        logger.error("FEC introuvable pour section 4 : %s", fec_path)
        lignes.append(f"⚠️  FEC introuvable : {fec_path}")
        return lignes

    try:
        from src.parsers.fec_parser import parse
        df = parse(str(fec_path))
    except Exception as exc:
        logger.error("Erreur parsing FEC (section 4) : %s", exc, exc_info=True)
        lignes.append(f"✗ Erreur parsing FEC : {exc}")
        return lignes

    # Agréger par compte (Solde net = Débit − Crédit, en K€)
    bal_ke = df.groupby("CompteNum")["Solde"].sum() / 1_000
    bal_ke.index = bal_ke.index.astype(str)

    def _sum_classes(*classes: str) -> float:
        return round(
            bal_ke[bal_ke.index.str[0].isin(set(classes))].sum(), 3
        )

    def _cmp(val: float, ref: float, tol: float = 0.1) -> str:
        ecart = abs(val - ref)
        return "✓" if ecart <= tol else f"⚠️  ÉCART {ecart:.3f} K€ > {tol}"

    # --- Résultat net : −cl7 − cl6 ---
    # cl.7 : soldes négatifs (crédités = produits), cl.6 : positifs (débités = charges).
    # Résultat = Produits − Charges = −Solde_KE_7 − Solde_KE_6
    s7 = _sum_classes("7")
    s6 = _sum_classes("6")
    resultat_ke  = round(-s7 - s6, 3)
    ref_r, tol_r = 2572.7, 0.1
    ok_r = _cmp(resultat_ke, ref_r, tol_r)
    lignes.append(
        f"- Résultat net théorique (−cl.7 − cl.6)"
        f" = **{resultat_ke:.3f} K€**"
        f" (attendu {ref_r} K€ ± {tol_r})"
        f" — {ok_r}"
    )
    if "⚠️" in ok_r:
        lignes.append(
            "  → Le FEC ou le pipeline a changé, reconduire l'analyse."
        )

    # --- Actif brut : classes 2+3+5 ---
    actif_ke = _sum_classes("2", "3", "5")
    ref_a, tol_a = 8790.5, 0.1
    ok_a = _cmp(actif_ke, ref_a, tol_a)
    lignes.append(
        f"- Actif brut (classes 2+3+5) = **{actif_ke:.3f} K€**"
        f" (attendu {ref_a} K€ ± {tol_a}) — {ok_a}"
    )
    if "⚠️" in ok_a:
        lignes.append(
            "  → Le FEC ou le pipeline a changé, reconduire l'analyse."
        )

    # --- Passif brut signé : classes 1+4 ---
    passif_ke = _sum_classes("1", "4")
    ref_p, tol_p = -6217.8, 0.1
    ok_p = _cmp(passif_ke, ref_p, tol_p)
    lignes.append(
        f"- Passif brut signé (classes 1+4) = **{passif_ke:.3f} K€**"
        f" (attendu {ref_p} K€ ± {tol_p}) — {ok_p}"
    )
    if "⚠️" in ok_p:
        lignes.append(
            "  → Le FEC ou le pipeline a changé, reconduire l'analyse."
        )

    # --- 4b. Résiduel bilan (depuis les WARNINGs capturés lors du pipeline) ---
    lignes.append("\n### 4b. Résiduel bilan (WARNING logger `fm_writer`)\n")
    residuels = [
        r for r in warnings_captures
        if "résiduel" in r.lower() or "residuel" in r.lower()
    ]
    if residuels:
        for r in residuels:
            lignes.append(f"- `{r}`")
            m = re.search(r"=\s*([-\d.]+)\s*K€", r)
            if m:
                val = float(m.group(1))
                ref_res, tol_res = 760.725, 0.1
                ecart_res = abs(abs(val) - ref_res)
                ok_res = (
                    "✓" if ecart_res <= tol_res
                    else f"⚠️  ÉCART {ecart_res:.3f} K€ > {tol_res}"
                )
                lignes.append(
                    f"  → Valeur extraite : {val:.3f} K€"
                    f" (attendu {ref_res} K€ ± {tol_res}) — {ok_res}"
                )
    else:
        lignes.append(
            "Aucun WARNING résiduel capturé"
            " (résiduel ≤ 0.1 K€ ou pipeline non exécuté)."
        )

    return lignes


# ---------------------------------------------------------------------------
# Section 5 — Tests
# ---------------------------------------------------------------------------
def _section5() -> List[str]:
    lignes = ["## 5. Tests\n"]

    # S'assurer que pytest est disponible
    r0 = subprocess.run(
        [sys.executable, "-m", "pytest", "--version"],
        capture_output=True, text=True, cwd=str(ROOT),
    )
    if r0.returncode != 0:
        logger.info("pytest non trouvé — installation en cours…")
        inst = subprocess.run(
            [sys.executable, "-m", "pip", "install", "pytest",
             "--break-system-packages"],
            capture_output=True, text=True,
        )
        if inst.returncode != 0:
            logger.error("Impossible d'installer pytest : %s", inst.stderr)
            lignes.append(
                f"✗ Installation pytest impossible : {inst.stderr[:300]}"
            )
            return lignes
        logger.info("pytest installé avec succès")

    result = subprocess.run(
        [sys.executable, "-m", "pytest", "tests/", "-v", "--tb=short"],
        capture_output=True, text=True, cwd=str(ROOT),
    )
    output_lines = (result.stdout + result.stderr).splitlines()

    # Extraire le résumé (dernière ligne du type "X passed, Y failed …")
    nb_passed = nb_failed = nb_error = 0
    for line in reversed(output_lines):
        if "passed" in line or "failed" in line or "error" in line:
            if m := re.search(r"(\d+) passed", line):
                nb_passed = int(m.group(1))
            if m := re.search(r"(\d+) failed", line):
                nb_failed = int(m.group(1))
            if m := re.search(r"(\d+) error", line):
                nb_error = int(m.group(1))
            break

    lignes.append(
        f"**Résultat** : {nb_passed} passé(s),"
        f" {nb_failed} échoué(s), {nb_error} en erreur\n"
    )

    # Lister les tests échoués avec une ligne de traceback
    echec_lines = [
        l.strip() for l in output_lines
        if " FAILED " in l or " ERROR " in l
    ]
    if echec_lines:
        lignes.append("**Tests échoués / en erreur :**\n")
        for nom_test in echec_lines:
            lignes.append(f"- `{nom_test}`")
            # Chercher une ligne d'erreur dans les 10 lignes suivantes
            try:
                raw_idx = next(
                    i for i, l in enumerate(output_lines)
                    if l.strip() == nom_test
                )
                for j in range(raw_idx + 1, min(raw_idx + 10, len(output_lines))):
                    l = output_lines[j]
                    if l.strip().startswith("E ") or "Error" in l or "AssertionError" in l:
                        lignes.append(f"  `{l.strip()[:120]}`")
                        break
            except StopIteration:
                pass
    elif nb_failed == 0 and nb_error == 0:
        lignes.append("Tous les tests passent ✓")

    # Sortie pytest brute (les 50 dernières lignes)
    lignes.append("\n<details><summary>Sortie pytest complète (fin)</summary>\n\n```")
    lignes.extend(output_lines[-50:])
    lignes.append("```\n</details>")

    return lignes


# ---------------------------------------------------------------------------
# Main — assemblage du rapport
# ---------------------------------------------------------------------------
def main() -> None:
    logger.info("=== Audit d'état du projet audit-automation-2 ===")

    output_dir = ROOT / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_tmp = ROOT / "output/audit_tmp"

    # -----------------------------------------------------------------------
    # Exécuter le pipeline UNE SEULE FOIS, capturer les WARNINGs
    # -----------------------------------------------------------------------
    fec_path   = ROOT / "data/GILAC_2025_12_31_FEC.txt"
    n1_fm_path = ROOT / "data/FM GILAC.xlsx"
    tpl_path   = ROOT / "data/templates"

    cap = _CapturingHandler(logging.WARNING)
    logging.getLogger().addHandler(cap)

    resultats: Optional[Dict] = None
    pipeline_error: Optional[Exception] = None

    if fec_path.exists():
        logger.info("Exécution du pipeline GILAC → output/audit_tmp/")
        try:
            from main import run_pipeline  # noqa: E402
            resultats = run_pipeline(
                fec_path=str(fec_path),
                client="GILAC",
                date_cloture="31/12/2025",
                n1_fm=str(n1_fm_path) if n1_fm_path.exists() else None,
                templates_dir=str(tpl_path) if tpl_path.exists() else None,
                output_dir=str(output_tmp),
            )
        except Exception as exc:
            logger.error("Erreur pipeline : %s", exc, exc_info=True)
            pipeline_error = exc
    else:
        logger.error("FEC GILAC introuvable : %s", fec_path)
        pipeline_error = FileNotFoundError(str(fec_path))

    logging.getLogger().removeHandler(cap)
    logger.info(
        "%d WARNINGs capturés durant le pipeline", len(cap.records)
    )

    # -----------------------------------------------------------------------
    # Assembler les sections
    # -----------------------------------------------------------------------
    today = date.today().strftime("%d/%m/%Y")
    sections: List[str] = [
        "# Rapport d'audit — État du projet audit-automation-2\n",
        f"**Date :** {today}  \n**Projet :** audit-automation-2\n",
        "---\n",
    ]

    logger.info("Génération section 1 — Fichiers attendus")
    sections += _section1()
    sections.append("\n---\n")

    logger.info("Génération section 2 — Livrables")
    sections += _section2(resultats, pipeline_error, output_tmp)
    sections.append("\n---\n")

    logger.info("Génération section 3 — Dette technique")
    sections += _section3()
    sections.append("\n---\n")

    logger.info("Génération section 4 — Cibles chiffrées GILAC")
    sections += _section4(cap.records)
    sections.append("\n---\n")

    logger.info("Génération section 5 — Tests")
    sections += _section5()

    rapport_path = output_dir / "AUDIT_ETAT.md"
    rapport_path.write_text("\n".join(sections), encoding="utf-8")
    logger.info("=== Rapport généré : %s ===", rapport_path)


if __name__ == "__main__":
    main()
